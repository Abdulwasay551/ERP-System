"""
Human-readable Excel rendering of a backup export (Phase B) - for a shop owner to open
and eyeball that a backup "looks right," NOT a restore path. No code anywhere reads an
.xlsx file back in; the JSON snapshot format from core.snapshot stays the sole
authoritative source for actually restoring data (see core.snapshot's own module
docstring for why - it round-trips through both Postgres and SQLite identically, an
Excel file can't).

Follows the same lazy-import + graceful-degradation convention already used for Excel
exports elsewhere in this codebase (see accounting/views.py) - openpyxl is a real
dependency (already in requirements.txt) so this should never actually hit the
ImportError path in practice, but callers still handle it the same way as that
existing convention, for consistency.
"""
from collections import defaultdict


def build_backup_workbook(objects_data):
    """Given the same {'model', 'pk', 'fields'} list core.snapshot's export functions
    produce, returns an openpyxl Workbook with one sheet per model - column headers
    from the union of field names seen for that model, one row per object, 'pk' as the
    first column. Raises ImportError if openpyxl isn't installed (caller's
    responsibility to catch, matching the rest of this codebase's Excel-export
    convention)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    by_model = defaultdict(list)
    for obj in objects_data:
        by_model[obj['model']].append(obj)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # the default blank sheet - each model gets its own below

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    for model_label in sorted(by_model.keys()):
        rows = by_model[model_label]
        # Excel sheet names: max 31 chars, no []:*?/\\
        sheet_name = model_label.replace('.', '_')[:31]
        ws = wb.create_sheet(sheet_name)

        field_names = []
        seen = set()
        for obj in rows:
            for key in obj['fields'].keys():
                if key not in seen:
                    seen.add(key)
                    field_names.append(key)
        headers = ['id'] + field_names

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_num, obj in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=obj['pk'])
            for col, field in enumerate(field_names, 2):
                value = obj['fields'].get(field)
                # openpyxl can't write lists/dicts (JSONField values, M2M lists) - stringify
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_num, column=col, value=value)

    if not wb.sheetnames:
        wb.create_sheet('Empty')  # a workbook needs at least one sheet
    return wb
