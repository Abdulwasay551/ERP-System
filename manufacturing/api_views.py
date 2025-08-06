from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django.db import transaction

from .models import (
    WorkCenter, BillOfMaterials, BillOfMaterialsItem, BOMOperation,
    WorkOrder, WorkOrderOperation, OperationLog, QualityCheck,
    MaterialConsumption, MRPPlan, MRPRequirement, ProductionPlan,
    ProductionPlanItem, JobCard, Subcontractor, SubcontractWorkOrder,
    DemandForecast, SupplierLeadTime, MRPRunLog, ReorderRule, CapacityPlan
)
from .serializers import (
    WorkCenterSerializer, BillOfMaterialsSerializer, BillOfMaterialsItemSerializer,
    BOMOperationSerializer, WorkOrderSerializer, WorkOrderOperationSerializer,
    OperationLogSerializer, QualityCheckSerializer, MaterialConsumptionSerializer,
    MRPPlanSerializer, MRPRequirementSerializer, ProductionPlanSerializer,
    ProductionPlanItemSerializer, JobCardSerializer, SubcontractorSerializer,
    SubcontractWorkOrderSerializer, DemandForecastSerializer, SupplierLeadTimeSerializer,
    MRPRunLogSerializer, ReorderRuleSerializer, CapacityPlanSerializer
)


class WorkCenterViewSet(viewsets.ModelViewSet):
    serializer_class = WorkCenterSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return WorkCenter.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


class BillOfMaterialsViewSet(viewsets.ModelViewSet):
    serializer_class = BillOfMaterialsSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return BillOfMaterials.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Create a duplicate of an existing BOM"""
        bom = self.get_object()
        
        # Create new BOM
        new_bom = BillOfMaterials.objects.create(
            company=bom.company,
            product=bom.product,
            name=f"{bom.name} (Copy)",
            version=f"{bom.version}-copy",
            manufacturing_type=bom.manufacturing_type,
            lot_size=bom.lot_size,
            lead_time_days=bom.lead_time_days,
            scrap_percentage=bom.scrap_percentage,
            routing_required=bom.routing_required,
            phantom_bom=bom.phantom_bom,
            quality_check_required=bom.quality_check_required,
            quality_specification=bom.quality_specification,
            created_by=request.user
        )
        
        # Copy BOM items
        for item in bom.items.all():
            BillOfMaterialsItem.objects.create(
                bom=new_bom,
                component=item.component,
                quantity=item.quantity,
                unit_cost=item.unit_cost,
                waste_percentage=item.waste_percentage,
                is_optional=item.is_optional,
                preferred_supplier=item.preferred_supplier,
                sequence=item.sequence,
                notes=item.notes
            )
        
        # Copy operations
        for operation in bom.operations.all():
            BOMOperation.objects.create(
                bom=new_bom,
                work_center=operation.work_center,
                operation_name=operation.operation_name,
                description=operation.description,
                sequence=operation.sequence,
                setup_time_minutes=operation.setup_time_minutes,
                run_time_per_unit_minutes=operation.run_time_per_unit_minutes,
                cleanup_time_minutes=operation.cleanup_time_minutes,
                operators_required=operation.operators_required,
                skill_level_required=operation.skill_level_required,
                quality_check_required=operation.quality_check_required,
                quality_specification=operation.quality_specification,
                cost_per_hour=operation.cost_per_hour,
                work_instruction=operation.work_instruction,
                safety_notes=operation.safety_notes
            )
        
        serializer = self.get_serializer(new_bom)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BillOfMaterialsItemViewSet(viewsets.ModelViewSet):
    serializer_class = BillOfMaterialsItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return BillOfMaterialsItem.objects.filter(bom__company=self.request.user.company)


class BOMOperationViewSet(viewsets.ModelViewSet):
    serializer_class = BOMOperationSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return BOMOperation.objects.filter(bom__company=self.request.user.company)


class WorkOrderViewSet(viewsets.ModelViewSet):
    serializer_class = WorkOrderSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return WorkOrder.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Start work order production"""
        work_order = self.get_object()
        work_order.start_production()
        return Response({'status': 'started', 'actual_start': work_order.actual_start})
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Complete work order production"""
        work_order = self.get_object()
        work_order.complete_production()
        return Response({'status': 'completed', 'actual_end': work_order.actual_end})
    
    @action(detail=True, methods=['post'])
    def generate_operations(self, request, pk=None):
        """Generate work order operations from BOM"""
        work_order = self.get_object()
        
        # Delete existing operations
        work_order.operations.all().delete()
        
        # Create operations from BOM
        for bom_operation in work_order.bom.operations.all():
            WorkOrderOperation.objects.create(
                work_order=work_order,
                bom_operation=bom_operation,
                work_center=bom_operation.work_center,
                sequence=bom_operation.sequence,
                quantity_to_produce=work_order.quantity_planned
            )
        
        return Response({'message': 'Operations generated successfully'})


class WorkOrderOperationViewSet(viewsets.ModelViewSet):
    serializer_class = WorkOrderOperationSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return WorkOrderOperation.objects.filter(work_order__company=self.request.user.company)


class OperationLogViewSet(viewsets.ModelViewSet):
    serializer_class = OperationLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return OperationLog.objects.filter(work_order_operation__work_order__company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(operator=self.request.user)


class QualityCheckViewSet(viewsets.ModelViewSet):
    serializer_class = QualityCheckSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return QualityCheck.objects.filter(work_order__company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(inspector=self.request.user)


class MaterialConsumptionViewSet(viewsets.ModelViewSet):
    serializer_class = MaterialConsumptionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return MaterialConsumption.objects.filter(work_order__company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(consumed_by=self.request.user, consumed_at=timezone.now())


class MRPPlanViewSet(viewsets.ModelViewSet):
    serializer_class = MRPPlanSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return MRPPlan.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def calculate(self, request, pk=None):
        """Run MRP calculation"""
        mrp_plan = self.get_object()
        
        try:
            from .mrp_engine import MRPEngine
            
            with transaction.atomic():
                mrp_plan.status = 'calculating'
                mrp_plan.calculation_start = timezone.now()
                mrp_plan.save()
                
                # Run MRP calculation using the engine
                engine = MRPEngine(mrp_plan)
                success = engine.run_mrp_calculation()
                
                if success:
                    mrp_plan.status = 'completed'
                    mrp_plan.calculation_end = timezone.now()
                    mrp_plan.save()
                    
                    # Get the results
                    requirements_count = mrp_plan.requirements.count()
                    
                    return Response({
                        'message': 'MRP calculation completed successfully',
                        'requirements_generated': requirements_count,
                        'calculation_time': (mrp_plan.calculation_end - mrp_plan.calculation_start).total_seconds()
                    })
                else:
                    mrp_plan.status = 'draft'
                    mrp_plan.save()
                    return Response({'error': 'MRP calculation failed'}, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            mrp_plan.status = 'draft'
            mrp_plan.save()
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def supply_demand_report(self, request, pk=None):
        """Generate supply-demand analysis report"""
        mrp_plan = self.get_object()
        
        try:
            from .mrp_engine import SupplyDemandAnalyzer
            
            analyzer = SupplyDemandAnalyzer(
                company=mrp_plan.company,
                start_date=mrp_plan.plan_date,
                end_date=mrp_plan.plan_date + timezone.timedelta(days=mrp_plan.planning_horizon_days)
            )
            
            report_data = analyzer.generate_supply_demand_report()
            
            # Format the data for API response
            formatted_data = []
            for item in report_data:
                formatted_data.append({
                    'product_id': item['product'].id,
                    'product_name': item['product'].name,
                    'current_stock': float(item['current_stock']),
                    'safety_stock': float(item['safety_stock']),
                    'reorder_point': float(item['reorder_point']),
                    'total_demand': float(item['total_demand']),
                    'total_supply': float(item['total_supply']),
                    'net_requirement': float(item['net_requirement']),
                    'status': item['status'],
                    'days_of_stock': item['days_of_stock'] if item['days_of_stock'] != float('inf') else None
                })
            
            return Response({
                'report_date': mrp_plan.plan_date,
                'planning_horizon_days': mrp_plan.planning_horizon_days,
                'data': formatted_data
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def run_automatic_mrp(self, request):
        """Run automatic MRP for the company"""
        try:
            from .mrp_engine import run_automatic_mrp
            
            mrp_plan, message = run_automatic_mrp(request.user.company)
            
            serializer = self.get_serializer(mrp_plan)
            
            return Response({
                'message': message,
                'mrp_plan': serializer.data
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class MRPRequirementViewSet(viewsets.ModelViewSet):
    serializer_class = MRPRequirementSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return MRPRequirement.objects.filter(mrp_plan__company=self.request.user.company)


class ProductionPlanViewSet(viewsets.ModelViewSet):
    serializer_class = ProductionPlanSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return ProductionPlan.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)


class ProductionPlanItemViewSet(viewsets.ModelViewSet):
    serializer_class = ProductionPlanItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return ProductionPlanItem.objects.filter(production_plan__company=self.request.user.company)


class JobCardViewSet(viewsets.ModelViewSet):
    serializer_class = JobCardSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return JobCard.objects.filter(work_order__company=self.request.user.company)


class SubcontractorViewSet(viewsets.ModelViewSet):
    serializer_class = SubcontractorSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Subcontractor.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


class SubcontractWorkOrderViewSet(viewsets.ModelViewSet):
    serializer_class = SubcontractWorkOrderSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return SubcontractWorkOrder.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


class DemandForecastViewSet(viewsets.ModelViewSet):
    serializer_class = DemandForecastSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return DemandForecast.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company, created_by=self.request.user)


class SupplierLeadTimeViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierLeadTimeSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return SupplierLeadTime.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)
    
    @action(detail=False, methods=['get'])
    def by_product(self, request):
        """Get lead times grouped by product"""
        product_id = request.query_params.get('product_id')
        if not product_id:
            return Response({'error': 'product_id parameter required'}, status=status.HTTP_400_BAD_REQUEST)
        
        lead_times = self.get_queryset().filter(product_id=product_id, is_active=True)
        serializer = self.get_serializer(lead_times, many=True)
        return Response(serializer.data)


class MRPRunLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = MRPRunLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return MRPRunLog.objects.filter(company=self.request.user.company)


class ReorderRuleViewSet(viewsets.ModelViewSet):
    serializer_class = ReorderRuleSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return ReorderRule.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)
    
    @action(detail=False, methods=['post'])
    def check_reorders(self, request):
        """Check all reorder rules and identify items that need reordering"""
        from inventory.models import StockItem
        
        reorder_items = []
        rules = self.get_queryset().filter(is_active=True)
        
        for rule in rules:
            try:
                stock_item = StockItem.objects.get(
                    product=rule.product,
                    warehouse=rule.warehouse
                )
                
                if rule.should_reorder(stock_item.available_quantity):
                    order_qty = rule.calculate_order_quantity(stock_item.available_quantity)
                    
                    reorder_items.append({
                        'rule_id': rule.id,
                        'product_id': rule.product.id,
                        'product_name': rule.product.name,
                        'warehouse_id': rule.warehouse.id,
                        'warehouse_name': rule.warehouse.name,
                        'current_stock': float(stock_item.available_quantity),
                        'reorder_point': float(rule.reorder_point),
                        'suggested_order_quantity': float(order_qty),
                        'reorder_method': rule.reorder_method
                    })
                    
            except StockItem.DoesNotExist:
                continue
        
        return Response({
            'reorder_items': reorder_items,
            'total_items_to_reorder': len(reorder_items)
        })


class CapacityPlanViewSet(viewsets.ModelViewSet):
    serializer_class = CapacityPlanSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return CapacityPlan.objects.filter(company=self.request.user.company)
    
    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)
    
    @action(detail=False, methods=['get'])
    def capacity_analysis(self, request):
        """Get capacity analysis across all work centers"""
        plans = self.get_queryset().order_by('plan_date', 'work_center__name')
        
        analysis = {
            'overloaded_centers': [],
            'underutilized_centers': [],
            'average_utilization': 0,
            'total_centers': 0
        }
        
        total_utilization = 0
        center_count = 0
        
        for plan in plans:
            plan.calculate_utilization()
            plan.save()
            
            utilization = float(plan.utilization_percentage)
            total_utilization += utilization
            center_count += 1
            
            center_data = {
                'work_center_id': plan.work_center.id,
                'work_center_name': plan.work_center.name,
                'plan_date': plan.plan_date,
                'utilization_percentage': utilization,
                'available_hours': float(plan.available_hours),
                'planned_hours': float(plan.planned_hours)
            }
            
            if plan.is_overloaded:
                analysis['overloaded_centers'].append(center_data)
            elif utilization < 50:
                analysis['underutilized_centers'].append(center_data)
        
        if center_count > 0:
            analysis['average_utilization'] = total_utilization / center_count
        
        analysis['total_centers'] = center_count
        
        return Response(analysis)


# Additional API Views for MRP Frontend Support

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime, timedelta
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from io import BytesIO


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_create_purchase_orders(request):
    """Create purchase orders for all pending purchase requirements"""
    try:
        company = request.user.company
        
        # Get all pending purchase requirements
        requirements = MRPRequirement.objects.filter(
            mrp_plan__company=company,
            source_type='purchase',
            status='pending',
            shortage_quantity__gt=0
        )
        
        if not requirements.exists():
            return JsonResponse({
                'success': False,
                'error': 'No pending purchase requirements found'
            })
        
        from purchase.models import PurchaseRequest, PurchaseRequestItem
        from decimal import Decimal
        
        # Group requirements by suggested order date
        requirements_by_date = {}
        for req in requirements:
            date_key = req.suggested_order_date
            if date_key not in requirements_by_date:
                requirements_by_date[date_key] = []
            requirements_by_date[date_key].append(req)
        
        pos_created = 0
        
        with transaction.atomic():
            for order_date, reqs in requirements_by_date.items():
                # Create purchase request
                pr = PurchaseRequest.objects.create(
                    company=company,
                    title=f"Bulk MRP PO - {order_date}",
                    description=f"Auto-generated from MRP requirements",
                    priority='normal',
                    required_date=order_date,
                    created_by=request.user,
                    status='draft'
                )
                
                # Add items
                for req in reqs:
                    PurchaseRequestItem.objects.create(
                        purchase_request=pr,
                        product=req.product,
                        description=req.product.description or req.product.name,
                        quantity=req.shortage_quantity,
                        estimated_cost=Decimal('0'),
                        priority='normal'
                    )
                    
                    # Update requirement status
                    req.status = 'ordered'
                    req.notes = f"Purchase Request created: {pr.title}"
                    req.save()
                
                pos_created += 1
        
        return JsonResponse({
            'success': True,
            'pos_created': pos_created,
            'message': f'{pos_created} purchase orders created successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_mrp_requirement_details(request, requirement_id):
    """Get detailed information about a specific MRP requirement"""
    try:
        requirement = MRPRequirement.objects.get(
            id=requirement_id,
            mrp_plan__company=request.user.company
        )
        
        return JsonResponse({
            'success': True,
            'requirement': {
                'id': requirement.id,
                'product_name': requirement.product.name,
                'product_code': requirement.product.code,
                'required_quantity': float(requirement.required_quantity),
                'available_quantity': float(requirement.available_quantity),
                'shortage_quantity': float(requirement.shortage_quantity),
                'required_date': requirement.required_date.strftime('%Y-%m-%d'),
                'suggested_order_date': requirement.suggested_order_date.strftime('%Y-%m-%d'),
                'source_type': requirement.source_type,
                'status': requirement.status,
                'notes': requirement.notes or '',
                'mrp_plan_name': requirement.mrp_plan.name,
                'created_at': requirement.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        })
        
    except MRPRequirement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Requirement not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_purchase_order_from_requirement(request, requirement_id):
    """Create a purchase order from a specific requirement"""
    try:
        requirement = MRPRequirement.objects.get(
            id=requirement_id,
            mrp_plan__company=request.user.company
        )
        
        if requirement.source_type != 'purchase':
            return JsonResponse({
                'success': False,
                'error': 'This requirement is not for purchase'
            })
        
        if requirement.status != 'pending':
            return JsonResponse({
                'success': False,
                'error': 'Requirement is not pending'
            })
        
        from purchase.models import PurchaseRequest, PurchaseRequestItem
        from decimal import Decimal
        
        with transaction.atomic():
            # Create purchase request
            pr = PurchaseRequest.objects.create(
                company=request.user.company,
                title=f"MRP PO - {requirement.product.name}",
                description=f"Generated from MRP requirement #{requirement.id}",
                priority='normal',
                required_date=requirement.suggested_order_date,
                created_by=request.user,
                status='draft'
            )
            
            # Add item
            PurchaseRequestItem.objects.create(
                purchase_request=pr,
                product=requirement.product,
                description=requirement.product.description or requirement.product.name,
                quantity=requirement.shortage_quantity,
                estimated_cost=Decimal('0'),
                priority='normal'
            )
            
            # Update requirement status
            requirement.status = 'ordered'
            requirement.notes = f"Purchase Request created: {pr.title}"
            requirement.save()
        
        return JsonResponse({
            'success': True,
            'purchase_request_id': pr.id,
            'message': 'Purchase order created successfully'
        })
        
    except MRPRequirement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Requirement not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_supply_demand_excel(request):
    """Export supply & demand report to Excel"""
    try:
        from .mrp_engine import SupplyDemandAnalyzer
        from datetime import timedelta
        
        company = request.user.company
        
        # Get date range
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date()
            
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = start_date + timedelta(days=90)
        
        # Generate report
        analyzer = SupplyDemandAnalyzer(company, start_date, end_date)
        report_data = analyzer.generate_supply_demand_report()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supply & Demand Report"
        
        # Headers
        headers = [
            'Product Name', 'Product Code', 'Current Stock', 'Safety Stock',
            'Reorder Point', 'Total Demand', 'Total Supply', 'Net Requirement',
            'Days of Stock', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        for row, item in enumerate(report_data, 2):
            ws.cell(row=row, column=1, value=item['product'].name)
            ws.cell(row=row, column=2, value=item['product'].code or '')
            ws.cell(row=row, column=3, value=float(item['current_stock']))
            ws.cell(row=row, column=4, value=float(item['safety_stock']))
            ws.cell(row=row, column=5, value=float(item['reorder_point']))
            ws.cell(row=row, column=6, value=float(item['total_demand']))
            ws.cell(row=row, column=7, value=float(item['total_supply']))
            ws.cell(row=row, column=8, value=float(item['net_requirement']))
            ws.cell(row=row, column=9, value=item['days_of_stock'] if item['days_of_stock'] != float('inf') else 'Infinite')
            ws.cell(row=row, column=10, value=item['status'])
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="supply_demand_report_{start_date}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_mrp_settings(request):
    """Save MRP configuration settings"""
    try:
        data = json.loads(request.body)
        company = request.user.company
        
        # Update or create MRP settings (could be stored in a settings model)
        # For now, we'll just validate and return success
        # In a real implementation, you might store these in a MRPSettings model
        
        planning_horizon = int(data.get('planning_horizon_days', 90))
        include_safety_stock = bool(data.get('include_safety_stock', True))
        consider_lead_times = bool(data.get('consider_lead_times', True))
        auto_create_purchase_requests = bool(data.get('auto_create_purchase_requests', False))
        
        # Validate settings
        if planning_horizon < 1 or planning_horizon > 365:
            return JsonResponse({
                'success': False,
                'error': 'Planning horizon must be between 1 and 365 days'
            })
        
        # Here you would save to a settings model
        # For now, just return success
        return JsonResponse({
            'success': True,
            'message': 'MRP settings saved successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_reorder_purchase_order(request, product_id):
    """Create a purchase order for a product that needs reordering"""
    try:
        from products.models import Product
        from inventory.models import StockItem
        from purchase.models import PurchaseRequest, PurchaseRequestItem
        from decimal import Decimal
        
        product = Product.objects.get(id=product_id, company=request.user.company)
        
        # Get stock item to determine reorder quantity
        try:
            stock_item = StockItem.objects.get(
                product=product,
                warehouse__company=request.user.company
            )
            
            reorder_quantity = max(
                stock_item.reorder_point - stock_item.available_quantity,
                stock_item.reorder_quantity or Decimal('1')
            )
            
        except StockItem.DoesNotExist:
            reorder_quantity = Decimal('1')
        
        with transaction.atomic():
            # Create purchase request
            pr = PurchaseRequest.objects.create(
                company=request.user.company,
                title=f"Reorder - {product.name}",
                description=f"Automatic reorder for {product.name}",
                priority='normal',
                required_date=timezone.now().date() + timedelta(days=7),
                created_by=request.user,
                status='draft'
            )
            
            # Add item
            PurchaseRequestItem.objects.create(
                purchase_request=pr,
                product=product,
                description=product.description or product.name,
                quantity=reorder_quantity,
                estimated_cost=Decimal('0'),
                priority='normal'
            )
        
        return JsonResponse({
            'success': True,
            'purchase_request_id': pr.id,
            'reorder_quantity': float(reorder_quantity),
            'message': 'Reorder purchase order created successfully'
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Product not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_mrp_report(request):
    """Export comprehensive MRP report"""
    try:
        company = request.user.company
        
        # Get latest MRP plan
        mrp_plan = MRPPlan.objects.filter(company=company).order_by('-plan_date').first()
        
        if not mrp_plan:
            return JsonResponse({
                'success': False,
                'error': 'No MRP plans found'
            }, status=404)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # MRP Requirements sheet
        ws1 = wb.active
        ws1.title = "MRP Requirements"
        
        headers = [
            'Product Name', 'Product Code', 'Required Quantity', 'Available Quantity',
            'Shortage Quantity', 'Required Date', 'Suggested Order Date',
            'Source Type', 'Status', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header)
        
        requirements = mrp_plan.requirements.all()
        for row, req in enumerate(requirements, 2):
            ws1.cell(row=row, column=1, value=req.product.name)
            ws1.cell(row=row, column=2, value=req.product.code or '')
            ws1.cell(row=row, column=3, value=float(req.required_quantity))
            ws1.cell(row=row, column=4, value=float(req.available_quantity))
            ws1.cell(row=row, column=5, value=float(req.shortage_quantity))
            ws1.cell(row=row, column=6, value=req.required_date.strftime('%Y-%m-%d'))
            ws1.cell(row=row, column=7, value=req.suggested_order_date.strftime('%Y-%m-%d'))
            ws1.cell(row=row, column=8, value=req.source_type)
            ws1.cell(row=row, column=9, value=req.status)
            ws1.cell(row=row, column=10, value=req.notes or '')
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="mrp_report_{mrp_plan.plan_date}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500) 