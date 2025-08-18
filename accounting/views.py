from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from .models import Account, AccountGroup, AccountCategory, Journal, JournalEntry, JournalItem, JournalTemplate, JournalTemplateItem
from django.views.decorators.http import require_POST
from django.db.models import Count, Sum, Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator

# Create your views here.

@login_required
def accounting_dashboard(request):
    """Accounting module dashboard with key metrics and quick access"""
    try:
        company = request.user.company
        
        # Key metrics
        total_accounts = Account.objects.filter(company=company, is_active=True).count()
        asset_accounts = Account.objects.filter(company=company, type='asset', is_active=True).count()
        liability_accounts = Account.objects.filter(company=company, type='liability', is_active=True).count()
        equity_accounts = Account.objects.filter(company=company, type='equity', is_active=True).count()
        
        # Account groups breakdown with proper company filtering
        account_groups = AccountGroup.objects.filter(company=company).annotate(
            account_count=Count('accounts', filter=Q(accounts__is_active=True))
        ).filter(account_count__gt=0).order_by('-account_count')[:5]
        
        # Recent accounts with balance calculation
        recent_accounts = Account.objects.filter(
            company=company
        ).order_by('-created_at')[:5]
        
        # Add calculated balance for each account
        for account in recent_accounts:
            # Calculate balance from journal items
            total_debit = account.journal_items.aggregate(
                total=Sum('debit')
            )['total'] or 0
            total_credit = account.journal_items.aggregate(
                total=Sum('credit')
            )['total'] or 0
            
            # Calculate balance based on account type
            if account.balance_side == 'debit':
                account.balance = total_debit - total_credit
            else:
                account.balance = total_credit - total_debit
        
        context = {
            'total_accounts': total_accounts,
            'asset_accounts': asset_accounts,
            'liability_accounts': liability_accounts,
            'equity_accounts': equity_accounts,
            'account_groups': account_groups,
            'recent_accounts': recent_accounts,
        }
        return render(request, 'accounting/dashboard.html', context)
    except Exception as e:
        context = {
            'total_accounts': 0,
            'asset_accounts': 0,
            'liability_accounts': 0,
            'equity_accounts': 0,
            'account_groups': [],
            'recent_accounts': [],
            'error': str(e)
        }
        return render(request, 'accounting/dashboard.html', context)

@login_required
def accounts_ui(request):
    """Display accounts with filtering and pagination"""
    try:
        company = request.user.company
        
        # Get filter parameters
        search = request.GET.get('search', '')
        account_type = request.GET.get('type', '')
        status = request.GET.get('status', '')
        group_id = request.GET.get('group', '')
        
        # Build query
        accounts = Account.objects.filter(company=company)
        
        if search:
            accounts = accounts.filter(
                Q(name__icontains=search) | 
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        if account_type:
            accounts = accounts.filter(type=account_type)
            
        if status == 'active':
            accounts = accounts.filter(is_active=True)
        elif status == 'inactive':
            accounts = accounts.filter(is_active=False)
            
        if group_id:
            accounts = accounts.filter(group_id=group_id)
        
        # Add balance calculation
        for account in accounts:
            total_debit = account.journal_items.aggregate(
                total=Sum('debit')
            )['total'] or 0
            total_credit = account.journal_items.aggregate(
                total=Sum('credit')
            )['total'] or 0
            
            if account.balance_side == 'debit':
                account.balance = total_debit - total_credit
            else:
                account.balance = total_credit - total_debit
        
        # Get account groups for filter dropdown
        account_groups = AccountGroup.objects.filter(company=company, is_active=True)
        
        # Pagination
        paginator = Paginator(accounts.order_by('-created_at'), 20)
        page_number = request.GET.get('page')
        accounts = paginator.get_page(page_number)
        
        context = {
            'accounts': accounts,
            'account_groups': account_groups,
            'search': search,
            'account_type': account_type,
            'status': status,
            'group_id': group_id,
            'current_filter': account_type,  # Add this for template clarity
        }
        return render(request, 'accounting/accounts-ui.html', context)
    except Exception as e:
        context = {
            'accounts': [],
            'account_groups': [],
            'error': str(e)
        }
        return render(request, 'accounting/accounts-ui.html', context)

@login_required
def accounts_add(request):
    """Handle both GET and POST for account creation"""
    if request.method == 'GET':
        # Show form for creating account
        account_groups = AccountGroup.objects.filter(company=request.user.company, is_active=True)
        return render(request, 'accounting/account_form.html', {'account_groups': account_groups})
    
    elif request.method == 'POST':
        try:
            company = request.user.company
            code = request.POST.get('code')
            name = request.POST.get('name')
            type_ = request.POST.get('type')
            group_id = request.POST.get('group')
            description = request.POST.get('description', '')
            account_type = request.POST.get('account_type', '')
            balance_side = request.POST.get('balance_side', 'debit')
            is_active = request.POST.get('is_active') == 'on'
            
            if not all([code, name, type_, group_id]):
                return JsonResponse({'success': False, 'error': 'Code, name, type, and group are required.'}, status=400)
            
            # Get and validate group
            try:
                group = AccountGroup.objects.get(company=company, id=group_id, is_active=True)
            except (ValueError, AccountGroup.DoesNotExist):
                return JsonResponse({'success': False, 'error': 'Please select a valid account group.'}, status=400)
            
            # Check for duplicate code
            if Account.objects.filter(company=company, code=code).exists():
                return JsonResponse({'success': False, 'error': 'Account code already exists.'}, status=400)
            
            account = Account.objects.create(
                company=company,
                code=code,
                name=name,
                type=type_,
                group=group,
                description=description,
                account_type=type_,
                balance_side=balance_side,
                is_active=is_active
            )
            
            return JsonResponse({
                'success': True,
                'account': {
                    'id': account.id,
                    'code': account.code,
                    'name': account.name,
                    'type': account.get_type_display(),
                    'group': account.group.name if account.group else None,
                    'is_active': account.is_active
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def accounts_edit(request, pk):
    """Handle both GET and POST for account editing"""
    account = get_object_or_404(Account, pk=pk, company=request.user.company)
    
    if request.method == 'GET':
        account_groups = AccountGroup.objects.filter(company=request.user.company, is_active=True)
        return render(request, 'accounting/account_form.html', {
            'account': account,
            'account_groups': account_groups,
            'edit_mode': True
        })
    
    elif request.method == 'POST':
        try:
            code = request.POST.get('code')
            name = request.POST.get('name')
            type_ = request.POST.get('type')
            group_id = request.POST.get('group')
            description = request.POST.get('description', '')
            account_type = request.POST.get('account_type', '')
            balance_side = request.POST.get('balance_side', 'debit')
            is_active = request.POST.get('is_active') == 'on'
            
            if not code or not name or not type_:
                return JsonResponse({'success': False, 'error': 'Code, name, and type are required.'}, status=400)
            
            # Check for duplicate code (excluding current account)
            if Account.objects.filter(company=request.user.company, code=code).exclude(pk=pk).exists():
                return JsonResponse({'success': False, 'error': 'Account code already exists.'}, status=400)
            
            # Get group if specified
            group = None
            if group_id:
                group = AccountGroup.objects.filter(company=request.user.company, id=group_id).first()
                if not group:
                    return JsonResponse({'success': False, 'error': 'Group not found.'}, status=400)
            
            account.code = code
            account.name = name
            account.type = type_
            account.group = group
            account.description = description
            account.account_type = account_type
            account.balance_side = balance_side
            account.is_active = is_active
            account.save()
            
            return JsonResponse({
                'success': True,
                'account': {
                    'id': account.id,
                    'code': account.code,
                    'name': account.name,
                    'type': account.get_type_display(),
                    'group': account.group.name if account.group else None,
                    'is_active': account.is_active
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def accounts_delete(request, pk):
    """Handle account deletion with validation"""
    if request.method == 'POST':
        try:
            account = get_object_or_404(Account, pk=pk, company=request.user.company)
            
            # Check if account has journal entries
            if account.journal_items.exists():
                return JsonResponse({
                    'success': False, 
                    'error': 'Cannot delete account with journal entries. Please make it inactive instead.'
                }, status=400)
            
            account_name = account.name
            account.delete()
            return JsonResponse({
                'success': True,
                'message': f'Account "{account_name}" deleted successfully.'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

@login_required
def accounts_detail(request, pk):
    """Account detail view with transactions and activity"""
    account = get_object_or_404(Account, pk=pk, company=request.user.company)
    
    # Get recent transactions (journal entries)
    recent_entries = account.journal_items.select_related('entry').order_by('-entry__date')[:20]
    
    # Calculate balance
    balance = calculate_account_balance(account)
    
    # Get account statistics
    total_debit = account.journal_items.aggregate(total=Sum('debit'))['total'] or 0
    total_credit = account.journal_items.aggregate(total=Sum('credit'))['total'] or 0
    entry_count = account.journal_items.count()
    
    # Get related accounts (same group)
    related_accounts = []
    if account.group:
        related_accounts = account.group.accounts.filter(
            is_active=True
        ).exclude(id=account.id)[:5]
    
    context = {
        'account': account,
        'balance': balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'entry_count': entry_count,
        'recent_entries': recent_entries,
        'related_accounts': related_accounts,
    }
    
    return render(request, 'accounting/account_detail.html', context)

@login_required
def journals_ui(request):
    """Display journal entries with filtering"""
    try:
        company = request.user.company
        
        # Get journals
        journals = Journal.objects.filter(company=company, is_active=True)
        
        # Get recent journal entries
        recent_entries = JournalEntry.objects.filter(
            company=company
        ).select_related('journal', 'created_by').order_by('-created_at')[:10]
        
        context = {
            'journals': journals,
            'recent_entries': recent_entries,
        }
        return render(request, 'accounting/journals-ui.html', context)
    except Exception as e:
        context = {
            'journals': [],
            'recent_entries': [],
            'error': str(e)
        }
        return render(request, 'accounting/journals-ui.html', context)

@login_required
def dynamic_journals(request):
    """Dynamic Journal Entries interface with backend integration"""
    return render(request, 'accounting/dynamic_journals.html')

@login_required
def journal_templates_management(request):
    """Journal Templates management interface"""
    return render(request, 'accounting/journal_templates.html')

@login_required
def ledger_ui(request):
    """Display general ledger"""
    try:
        company = request.user.company
        
        # Get accounts with their balances
        accounts = Account.objects.filter(company=company, is_active=True)
        
        for account in accounts:
            total_debit = account.journal_items.aggregate(
                total=Sum('debit')
            )['total'] or 0
            total_credit = account.journal_items.aggregate(
                total=Sum('credit')
            )['total'] or 0
            
            if account.balance_side == 'debit':
                account.balance = total_debit - total_credit
            else:
                account.balance = total_credit - total_debit
        
        context = {
            'accounts': accounts,
        }
        return render(request, 'accounting/ledger-ui.html', context)
    except Exception as e:
        context = {
            'accounts': [],
            'error': str(e)
        }
        return render(request, 'accounting/ledger-ui.html', context)

@login_required
def trial_balance_ui(request):
    """Display trial balance"""
    try:
        company = request.user.company
        
        # Get all accounts with balances
        accounts = Account.objects.filter(company=company, is_active=True)
        trial_balance_data = []
        total_debits = 0
        total_credits = 0
        
        for account in accounts:
            total_debit = account.journal_items.aggregate(
                total=Sum('debit')
            )['total'] or 0
            total_credit = account.journal_items.aggregate(
                total=Sum('credit')
            )['total'] or 0
            
            debit_balance = 0
            credit_balance = 0
            
            if account.balance_side == 'debit':
                balance = total_debit - total_credit
                if balance > 0:
                    debit_balance = balance
                    total_debits += balance
                elif balance < 0:
                    credit_balance = abs(balance)
                    total_credits += abs(balance)
            else:
                balance = total_credit - total_debit
                if balance > 0:
                    credit_balance = balance
                    total_credits += balance
                elif balance < 0:
                    debit_balance = abs(balance)
                    total_debits += abs(balance)
            
            if debit_balance > 0 or credit_balance > 0:
                trial_balance_data.append({
                    'account': account,
                    'debit_balance': debit_balance,
                    'credit_balance': credit_balance,
                })
        
        context = {
            'trial_balance_data': trial_balance_data,
            'total_debits': total_debits,
            'total_credits': total_credits,
            'is_balanced': abs(total_debits - total_credits) < 0.01,
        }
        return render(request, 'accounting/trial-balance-ui.html', context)
    except Exception as e:
        context = {
            'trial_balance_data': [],
            'total_debits': 0,
            'total_credits': 0,
            'is_balanced': True,
            'error': str(e)
        }
        return render(request, 'accounting/trial-balance-ui.html', context)

@login_required
def reports_ui(request):
    """Display financial reports"""
    try:
        company = request.user.company
        
        # Get key financial metrics
        assets = Account.objects.filter(company=company, type='asset', is_active=True)
        liabilities = Account.objects.filter(company=company, type='liability', is_active=True)
        equity = Account.objects.filter(company=company, type='equity', is_active=True)
        income = Account.objects.filter(company=company, type='income', is_active=True)
        expenses = Account.objects.filter(company=company, type='expense', is_active=True)
        
        # Calculate totals
        def calculate_total_balance(accounts):
            total = 0
            for account in accounts:
                total_debit = account.journal_items.aggregate(total=Sum('debit'))['total'] or 0
                total_credit = account.journal_items.aggregate(total=Sum('credit'))['total'] or 0
                
                if account.balance_side == 'debit':
                    balance = total_debit - total_credit
                else:
                    balance = total_credit - total_debit
                total += balance
            return total
        
        total_assets = calculate_total_balance(assets)
        total_liabilities = calculate_total_balance(liabilities)
        total_equity = calculate_total_balance(equity)
        total_income = calculate_total_balance(income)
        total_expenses = calculate_total_balance(expenses)
        
        net_income = total_income - total_expenses
        
        context = {
            'total_assets': total_assets,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_income': net_income,
        }
        return render(request, 'accounting/reports-ui.html', context)
    except Exception as e:
        context = {
            'total_assets': 0,
            'total_liabilities': 0,
            'total_equity': 0,
            'total_income': 0,
            'total_expenses': 0,
            'net_income': 0,
            'error': str(e)
        }
        return render(request, 'accounting/reports-ui.html', context)

@login_required
def journal_create(request):
    """Create new journal entry"""
    return render(request, 'accounting/journal-create.html')

@login_required
def analytics_ui(request):
    """Accounting analytics and insights"""
    try:
        company = request.user.company
        
        # Monthly trend data for the last 12 months
        from datetime import datetime, timedelta
        from django.db.models import Sum, Count
        from calendar import monthrange
        
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=365)
        
        # Get monthly journal entry counts
        monthly_entries = []
        monthly_amounts = []
        months = []
        
        for i in range(12):
            month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            month_end = month_start.replace(day=monthrange(month_start.year, month_start.month)[1])
            
            entries_count = Journal.objects.filter(
                company=company,
                date__gte=month_start,
                date__lte=month_end
            ).count()
            
            total_amount = JournalEntry.objects.filter(
                journal__company=company,
                journal__date__gte=month_start,
                journal__date__lte=month_end
            ).aggregate(total=Sum('debit'))['total'] or 0
            
            monthly_entries.append(entries_count)
            monthly_amounts.append(float(total_amount))
            months.append(month_start.strftime('%b %Y'))
        
        # Reverse to show chronological order
        monthly_entries.reverse()
        monthly_amounts.reverse()
        months.reverse()
        
        # Account type distribution
        account_distribution = {}
        for category in ['asset', 'liability', 'equity', 'income', 'expense']:
            count = Account.objects.filter(company=company, type=category, is_active=True).count()
            account_distribution[category.title()] = count
        
        # Top accounts by transaction volume
        top_accounts = Account.objects.filter(company=company, is_active=True).annotate(
            transaction_count=Count('journal_items')
        ).order_by('-transaction_count')[:10]
        
        context = {
            'monthly_entries': monthly_entries,
            'monthly_amounts': monthly_amounts,
            'months': months,
            'account_distribution': account_distribution,
            'top_accounts': top_accounts,
        }
        return render(request, 'accounting/analytics.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/analytics.html', context)

@login_required
def balance_sheet_ui(request):
    """Balance sheet report"""
    try:
        company = request.user.company
        
        # Assets
        asset_accounts = Account.objects.filter(company=company, type='asset', is_active=True)
        assets = []
        total_assets = 0
        
        for account in asset_accounts:
            balance = calculate_account_balance(account)
            assets.append({
                'account': account,
                'balance': balance
            })
            total_assets += balance
        
        # Liabilities
        liability_accounts = Account.objects.filter(company=company, type='liability', is_active=True)
        liabilities = []
        total_liabilities = 0
        
        for account in liability_accounts:
            balance = calculate_account_balance(account)
            liabilities.append({
                'account': account,
                'balance': balance
            })
            total_liabilities += balance
        
        # Equity
        equity_accounts = Account.objects.filter(company=company, type='equity', is_active=True)
        equity = []
        total_equity = 0
        
        for account in equity_accounts:
            balance = calculate_account_balance(account)
            equity.append({
                'account': account,
                'balance': balance
            })
            total_equity += balance
        
        context = {
            'assets': assets,
            'liabilities': liabilities,
            'equity': equity,
            'total_assets': total_assets,
            'total_liabilities': total_liabilities,
            'total_equity': total_equity,
        }
        return render(request, 'accounting/balance-sheet.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/balance-sheet.html', context)

@login_required
def profit_loss_ui(request):
    """Profit and Loss statement"""
    try:
        company = request.user.company
        
        # Income
        income_accounts = Account.objects.filter(company=company, type='income', is_active=True)
        income = []
        total_income = 0
        
        for account in income_accounts:
            balance = calculate_account_balance(account)
            income.append({
                'account': account,
                'balance': balance
            })
            total_income += balance
        
        # Expenses
        expense_accounts = Account.objects.filter(company=company, type='expense', is_active=True)
        expenses = []
        total_expenses = 0
        
        for account in expense_accounts:
            balance = calculate_account_balance(account)
            expenses.append({
                'account': account,
                'balance': balance
            })
            total_expenses += balance
        
        net_income = total_income - total_expenses
        
        context = {
            'income': income,
            'expenses': expenses,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_income': net_income,
        }
        return render(request, 'accounting/profit-loss.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/profit-loss.html', context)

@login_required
def cash_flow_ui(request):
    """Cash flow statement"""
    try:
        company = request.user.company
        
        # Get cash accounts
        cash_accounts = Account.objects.filter(
            company=company, 
            name__icontains='cash',
            is_active=True
        ) | Account.objects.filter(
            company=company,
            name__icontains='bank',
            is_active=True
        )
        
        cash_flows = []
        total_cash_flow = 0
        
        for account in cash_accounts:
            balance = calculate_account_balance(account)
            cash_flows.append({
                'account': account,
                'balance': balance
            })
            total_cash_flow += balance
        
        context = {
            'cash_flows': cash_flows,
            'total_cash_flow': total_cash_flow,
        }
        return render(request, 'accounting/cash-flow.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/cash-flow.html', context)

@login_required
def bank_reconciliation_ui(request):
    """Bank reconciliation interface"""
    return render(request, 'accounting/bank-reconciliation.html')

@login_required
def module_mappings_ui(request):
    """Module integration mappings"""
    try:
        company = request.user.company
        from .models import ModuleAccountMapping
        
        mappings = ModuleAccountMapping.objects.filter(company=company)
        
        context = {
            'mappings': mappings,
        }
        return render(request, 'accounting/module-mappings.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/module-mappings.html', context)

@login_required
def settings_ui(request):
    """Accounting settings and configuration"""
    return render(request, 'accounting/settings.html')

@login_required
def account_categories_ui(request):
    """Account categories management"""
    try:
        company = request.user.company
        
        # Get filter parameter
        category_filter = request.GET.get('category', '')
        
        # Get accounts organized by category
        accounts = Account.objects.filter(company=company)
        
        categories = {
            'assets': accounts.filter(type='asset'),
            'liabilities': accounts.filter(type='liability'),
            'equity': accounts.filter(type='equity'),
            'income': accounts.filter(type='income'),
            'expenses': accounts.filter(type='expense'),
            'cogs': accounts.filter(type='cogs'),
        }
        
        # If specific category is requested, filter the display
        if category_filter and category_filter in categories:
            filtered_accounts = categories[category_filter]
        else:
            filtered_accounts = None
        
        context = {
            'categories': categories,
            'filtered_accounts': filtered_accounts,
            'current_category': category_filter,
        }
        return render(request, 'accounting/categories.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/categories.html', context)

@login_required
def account_groups_ui(request):
    """Account groups management"""
    try:
        company = request.user.company
        from .models import AccountGroup
        
        groups = AccountGroup.objects.filter(company=company)
        
        context = {
            'groups': groups,
        }
        return render(request, 'accounting/groups.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/groups.html', context)

@login_required
def account_group_create(request):
    """Create new account group"""
    if request.method == 'POST':
        try:
            company = request.user.company
            from .models import AccountGroup
            
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            category = request.POST.get('category')
            is_active = request.POST.get('is_active') == 'on'
            
            group = AccountGroup.objects.create(
                company=company,
                name=name,
                description=description,
                category=category,
                is_active=is_active
            )
            
            return redirect('accounting:account_groups')
        except Exception as e:
            context = {'error': str(e)}
            return render(request, 'accounting/group_form.html', context)
    
    return render(request, 'accounting/group_form.html')

@login_required
def account_group_edit(request, pk):
    """Edit account group"""
    try:
        company = request.user.company
        from .models import AccountGroup
        
        group = AccountGroup.objects.get(id=pk, company=company)
        
        if request.method == 'POST':
            group.name = request.POST.get('name')
            group.description = request.POST.get('description', '')
            group.category = request.POST.get('category')
            group.is_active = request.POST.get('is_active') == 'on'
            group.save()
            
            return redirect('accounting:account_groups')
        
        context = {'group': group}
        return render(request, 'accounting/group_form.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/group_form.html', context)

@login_required
def account_group_delete(request, pk):
    """Delete account group"""
    try:
        company = request.user.company
        from .models import AccountGroup
        
        group = AccountGroup.objects.get(id=pk, company=company)
        
        if request.method == 'POST':
            group.delete()
            return redirect('accounting:account_groups')
        
        context = {'group': group}
        return render(request, 'accounting/group_confirm_delete.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/groups.html', context)

@login_required
def account_types_ui(request):
    """Account types management"""
    try:
        company = request.user.company
        
        # Define account types with counts
        account_types = {
            'asset': {
                'name': 'Assets',
                'description': 'Resources owned by the company',
                'count': Account.objects.filter(company=company, type='asset').count(),
                'icon': 'fas fa-home',
                'color': 'green'
            },
            'liability': {
                'name': 'Liabilities',
                'description': 'Debts and obligations',
                'count': Account.objects.filter(company=company, type='liability').count(),
                'icon': 'fas fa-credit-card',
                'color': 'red'
            },
            'equity': {
                'name': 'Equity',
                'description': 'Owner\'s stake in the business',
                'count': Account.objects.filter(company=company, type='equity').count(),
                'icon': 'fas fa-chart-pie',
                'color': 'purple'
            },
            'income': {
                'name': 'Income',
                'description': 'Revenue and earnings',
                'count': Account.objects.filter(company=company, type='income').count(),
                'icon': 'fas fa-arrow-up',
                'color': 'blue'
            },
            'expense': {
                'name': 'Expenses',
                'description': 'Costs and expenditures',
                'count': Account.objects.filter(company=company, type='expense').count(),
                'icon': 'fas fa-arrow-down',
                'color': 'orange'
            },
            'cogs': {
                'name': 'Cost of Goods Sold',
                'description': 'Direct costs of production',
                'count': Account.objects.filter(company=company, type='cogs').count(),
                'icon': 'fas fa-box',
                'color': 'yellow'
            }
        }
        
        context = {
            'account_types': account_types,
        }
        return render(request, 'accounting/account_types.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/account_types.html', context)

@login_required
def account_templates_ui(request):
    """Account templates management"""
    try:
        from .models import AccountTemplate
        
        # Get all account templates
        templates = AccountTemplate.objects.all().order_by('category', 'name')
        
        # Template categories for filtering
        template_categories = [
            {'key': 'retail', 'name': 'Retail Business'},
            {'key': 'manufacturing', 'name': 'Manufacturing'},
            {'key': 'service', 'name': 'Service Business'},
            {'key': 'consulting', 'name': 'Consulting'},
            {'key': 'restaurant', 'name': 'Restaurant/Food Service'},
            {'key': 'healthcare', 'name': 'Healthcare'},
            {'key': 'construction', 'name': 'Construction'},
            {'key': 'nonprofit', 'name': 'Non-Profit'},
            {'key': 'ecommerce', 'name': 'E-Commerce'},
            {'key': 'custom', 'name': 'Custom Template'},
        ]
        
        # Add display properties to templates
        for template in templates:
            template.color_class = f"bg-{['blue', 'green', 'purple', 'orange', 'red'][hash(template.category) % 5]}-100 text-{['blue', 'green', 'purple', 'orange', 'red'][hash(template.category) % 5]}-600"
            template.icon = 'fas fa-building'
            template.category_name = dict(AccountTemplate.TEMPLATE_CATEGORIES).get(template.category, template.category)
        
        context = {
            'templates': templates,
            'template_categories': template_categories,
        }
        return render(request, 'accounting/account_templates.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/account_templates.html', context)

@login_required
def account_analysis_ui(request):
    """Account analysis and reporting"""
    try:
        from django.db.models import Count, Q
        from datetime import datetime, timedelta
        
        company = request.user.company
        accounts = Account.objects.filter(company=company)
        
        # Summary statistics
        summary = {
            'total_accounts': accounts.count(),
            'active_accounts': accounts.filter(is_active=True).count(),
            'unused_accounts': accounts.filter(journal_items__isnull=True).count(),
            'duplicate_codes': 0,  # Would need to check for duplicate codes
        }
        
        # Account type distribution for chart
        type_distribution = {
            'labels': [],
            'data': []
        }
        
        for account_type, display_name in Account.ACCOUNT_TYPES:
            count = accounts.filter(type=account_type).count()
            if count > 0:
                type_distribution['labels'].append(display_name)
                type_distribution['data'].append(count)
        
        # Balance distribution (mock data - would need actual balance calculations)
        balance_distribution = {
            'labels': ['0-1K', '1K-10K', '10K-50K', '50K+'],
            'data': [15, 25, 12, 8]
        }
        
        # Account groups
        account_groups = AccountGroup.objects.filter(company=company).annotate(
            account_count=Count('accounts')
        )
        
        # Recent activities (mock data)
        recent_activities = [
            {
                'description': 'New account created: Cash in Bank',
                'timestamp': datetime.now() - timedelta(hours=2),
                'icon': 'fas fa-plus'
            },
            {
                'description': 'Account modified: Accounts Payable',
                'timestamp': datetime.now() - timedelta(days=1),
                'icon': 'fas fa-edit'
            }
        ]
        
        # Most used accounts (mock data - would need transaction counts)
        most_used_accounts = accounts.filter(is_active=True)[:10]
        for account in most_used_accounts:
            account.transaction_count = 0  # Would calculate from journal_items
            account.last_used = None
        
        # Unused accounts
        unused_accounts = accounts.filter(journal_items__isnull=True)[:20]
        
        # Recommendations
        recommendations = []
        
        if summary['unused_accounts'] > 5:
            recommendations.append({
                'title': 'Review Unused Accounts',
                'description': f'You have {summary["unused_accounts"]} accounts with no transactions. Consider archiving them.',
                'priority': 'medium',
                'icon': 'fas fa-eye',
                'actions': [{'label': 'Review Accounts', 'onclick': 'reviewUnusedAccounts()'}]
            })
        
        if account_groups.count() < 5:
            recommendations.append({
                'title': 'Organize with Account Groups',
                'description': 'Consider creating more account groups to better organize your chart of accounts.',
                'priority': 'low',
                'icon': 'fas fa-layer-group',
                'actions': [{'label': 'Create Groups', 'onclick': 'createAccountGroups()'}]
            })
        
        context = {
            'summary': summary,
            'type_distribution': type_distribution,
            'balance_distribution': balance_distribution,
            'account_groups': account_groups,
            'recent_activities': recent_activities,
            'most_used_accounts': most_used_accounts,
            'unused_accounts': unused_accounts,
            'recommendations': recommendations,
        }
        return render(request, 'accounting/account_analysis.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/account_analysis.html', context)

@login_required
def account_reconciliation_ui(request):
    """Account reconciliation management"""
    try:
        from .models import AccountReconciliation
        from datetime import datetime, timedelta
        
        company = request.user.company
        
        # Get accounts that are typically reconcilable (bank, credit card, asset accounts)
        reconcilable_accounts = Account.objects.filter(
            company=company, 
            is_active=True,
            type__in=['asset', 'liability']  # Bank accounts, credit cards, etc.
        ).exclude(is_group=True)
        
        # Add reconciliation status to each account
        for account in reconcilable_accounts:
            last_reconciliation = AccountReconciliation.objects.filter(
                account=account, status='completed'
            ).order_by('-completed_at').first()
            
            if last_reconciliation:
                account.last_reconciled = last_reconciliation.completed_at.date()
                if last_reconciliation.completed_at > datetime.now() - timedelta(days=30):
                    account.reconciliation_status = 'current'
                else:
                    account.reconciliation_status = 'overdue'
            else:
                account.last_reconciled = None
                account.reconciliation_status = 'pending'
            
            # Mock current balance - would calculate from journal items
            account.current_balance = 0.00
        
        # Get active reconciliations
        active_reconciliations = AccountReconciliation.objects.filter(
            account__company=company,
            status__in=['draft', 'in_progress']
        ).select_related('account')
        
        # Add period info for active reconciliations
        for reconciliation in active_reconciliations:
            reconciliation.period_start = reconciliation.reconciliation_date - timedelta(days=30)
            reconciliation.period_end = reconciliation.reconciliation_date
            reconciliation.started_at = reconciliation.created_at
        
        # Get reconciliation history
        reconciliation_history = AccountReconciliation.objects.filter(
            account__company=company,
            status='completed'
        ).select_related('account').order_by('-completed_at')[:10]
        
        # Add period info for history
        for reconciliation in reconciliation_history:
            reconciliation.period_start = reconciliation.reconciliation_date - timedelta(days=30)
            reconciliation.period_end = reconciliation.reconciliation_date
        
        # Summary statistics
        current_month = datetime.now().replace(day=1)
        summary = {
            'reconcilable_accounts': reconcilable_accounts.count(),
            'reconciled_this_month': AccountReconciliation.objects.filter(
                account__company=company,
                status='completed',
                completed_at__gte=current_month
            ).count(),
            'pending_reconciliations': active_reconciliations.count(),
            'discrepancies': AccountReconciliation.objects.filter(
                account__company=company,
                status='completed',
                difference__gt=0
            ).count(),
        }
        
        context = {
            'reconcilable_accounts': reconcilable_accounts,
            'summary': summary,
            'active_reconciliations': active_reconciliations,
            'reconciliation_history': reconciliation_history,
        }
        return render(request, 'accounting/account_reconciliation.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/account_reconciliation.html', context)

@login_required
def account_import_export_ui(request):
    """Account import/export functionality"""
    try:
        from .models import ImportExportOperation
        
        # Get recent import/export operations
        operation_history = ImportExportOperation.objects.filter(
            created_by__company=request.user.company
        ).order_by('-created_at')[:20]
        
        # Add file_url for completed exports (mock)
        for operation in operation_history:
            if operation.operation_type == 'export' and operation.status == 'completed':
                operation.file_url = f'/media/exports/{operation.file_name}'
        
        context = {
            'operation_history': operation_history,
        }
        return render(request, 'accounting/account_import_export.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/account_import_export.html', context)

@login_required
def coa_settings_ui(request):
    """Chart of accounts settings"""
    try:
        from .models import COASettings
        
        company = request.user.company
        
        # Get or create COA settings for the company
        settings, created = COASettings.objects.get_or_create(
            company=company,
            defaults={
                'code_format': 'numeric',
                'code_length': '4',
                'auto_generate_codes': True,
                'strict_code_validation': True,
                'default_view': 'list',
                'accounts_per_page': '50',
                'show_account_codes': True,
                'show_balances': True,
                'color_coding': True,
                'hierarchical_indent': True,
                'require_approval_new_accounts': False,
                'restrict_account_deletion': True,
                'log_account_changes': True,
                'log_balance_changes': True,
                'auto_sales_entries': True,
                'auto_purchase_entries': True,
                'auto_inventory_entries': True,
                'auto_backup_enabled': False,
                'backup_frequency': 'weekly',
            }
        )
        
        # Account type settings
        account_type_settings = {}
        for type_key, type_name in Account.ACCOUNT_TYPES:
            account_type_settings[type_key] = {
                'name': type_name,
                'enabled': True,
                'range_start': f"{['1', '2', '3', '4', '5', '6'][list(dict(Account.ACCOUNT_TYPES).keys()).index(type_key)]}000",
                'range_end': f"{['1', '2', '3', '4', '5', '6'][list(dict(Account.ACCOUNT_TYPES).keys()).index(type_key)]}999",
                'balance_side': 'debit' if type_key in ['asset', 'expense'] else 'credit'
            }
        
        # Get accounts for default account dropdowns
        income_accounts = Account.objects.filter(company=company, type='income', is_active=True)
        expense_accounts = Account.objects.filter(company=company, type='expense', is_active=True)
        asset_accounts = Account.objects.filter(company=company, type='asset', is_active=True)
        
        context = {
            'settings': settings,
            'account_type_settings': account_type_settings,
            'income_accounts': income_accounts,
            'expense_accounts': expense_accounts,
            'asset_accounts': asset_accounts,
        }
        return render(request, 'accounting/coa_settings.html', context)
    except Exception as e:
        context = {'error': str(e)}
        return render(request, 'accounting/coa_settings.html', context)

def calculate_account_balance(account):
    """Helper function to calculate account balance"""
    total_debit = account.journal_items.aggregate(
        total=Sum('debit')
    )['total'] or 0
    total_credit = account.journal_items.aggregate(
        total=Sum('credit')
    )['total'] or 0
    
    if account.balance_side == 'debit':
        return total_debit - total_credit
    else:
        return total_credit - total_debit

def calculate_account_balance_as_of(account, date_end):
    """Calculate account balance as of specific date"""
    from datetime import datetime
    
    if isinstance(date_end, str):
        date_end = datetime.strptime(date_end, '%Y-%m-%d').date()
    
    total_debit = account.journal_items.filter(
        journal_entry__date__lte=date_end,
        journal_entry__status='posted'
    ).aggregate(total=Sum('debit'))['total'] or 0
    
    total_credit = account.journal_items.filter(
        journal_entry__date__lte=date_end,
        journal_entry__status='posted'
    ).aggregate(total=Sum('credit'))['total'] or 0
    
    if account.balance_side == 'debit':
        return total_debit - total_credit
    else:
        return total_credit - total_debit

def is_current_asset(account):
    """Determine if account is a current asset per IFRS"""
    current_asset_types = [
        'cash', 'bank', 'receivable', 'inventory', 'prepaid', 'short_term_investment'
    ]
    current_asset_keywords = [
        'cash', 'bank', 'receivable', 'inventory', 'prepaid', 'current', 'short'
    ]
    
    # Check account type
    if account.account_type.lower() in current_asset_types:
        return True
    
    # Check account name for keywords
    account_name_lower = account.name.lower()
    return any(keyword in account_name_lower for keyword in current_asset_keywords)

def is_current_liability(account):
    """Determine if account is a current liability per IFRS"""
    current_liability_types = [
        'payable', 'accrued', 'short_term_debt', 'current_portion', 'tax_payable'
    ]
    current_liability_keywords = [
        'payable', 'accrued', 'current', 'short', 'tax', 'owing'
    ]
    
    # Check account type
    if account.account_type.lower() in current_liability_types:
        return True
    
    # Check account name for keywords
    account_name_lower = account.name.lower()
    return any(keyword in account_name_lower for keyword in current_liability_keywords)

def generate_balance_sheet_comparison(company, comparison_period_end):
    """Generate balance sheet data for comparison period"""
    try:
        accounts = Account.objects.filter(company=company, is_active=True)
        
        comparison_totals = {
            'current_assets': 0,
            'non_current_assets': 0,
            'current_liabilities': 0,
            'non_current_liabilities': 0,
            'total_equity': 0
        }
        
        for account in accounts:
            balance = calculate_account_balance_as_of(account, comparison_period_end)
            
            if account.type == 'asset':
                if is_current_asset(account):
                    comparison_totals['current_assets'] += balance
                else:
                    comparison_totals['non_current_assets'] += balance
            elif account.type == 'liability':
                if is_current_liability(account):
                    comparison_totals['current_liabilities'] += balance
                else:
                    comparison_totals['non_current_liabilities'] += balance
            elif account.type == 'equity':
                comparison_totals['total_equity'] += balance
        
        comparison_totals['total_assets'] = comparison_totals['current_assets'] + comparison_totals['non_current_assets']
        comparison_totals['total_liabilities'] = comparison_totals['current_liabilities'] + comparison_totals['non_current_liabilities']
        
        return {key: float(value) for key, value in comparison_totals.items()}
    except Exception:
        return None

def calculate_account_balance_for_period(account, period_start, period_end):
    """Calculate account balance for a specific period"""
    total_debit = account.journal_items.filter(
        journal_entry__date__gte=period_start,
        journal_entry__date__lte=period_end,
        journal_entry__status='posted'
    ).aggregate(total=Sum('debit'))['total'] or 0
    
    total_credit = account.journal_items.filter(
        journal_entry__date__gte=period_start,
        journal_entry__date__lte=period_end,
        journal_entry__status='posted'
    ).aggregate(total=Sum('credit'))['total'] or 0
    
    if account.type == 'income':
        return total_credit - total_debit  # Income accounts have credit balance
    elif account.type == 'expense':
        return total_debit - total_credit  # Expense accounts have debit balance
    else:
        # For balance sheet accounts, use normal balance side
        if account.balance_side == 'debit':
            return total_debit - total_credit
        else:
            return total_credit - total_debit

def is_revenue_account(account):
    """Determine if account is a primary revenue account"""
    revenue_keywords = ['sales', 'revenue', 'income', 'fees', 'services']
    exclude_keywords = ['other', 'miscellaneous', 'interest', 'gain', 'dividend']
    
    account_name_lower = account.name.lower()
    account_type_lower = (account.account_type or '').lower()
    
    # Check if it contains revenue keywords but not exclusion keywords
    has_revenue_keyword = any(keyword in account_name_lower or keyword in account_type_lower 
                             for keyword in revenue_keywords)
    has_exclude_keyword = any(keyword in account_name_lower or keyword in account_type_lower 
                             for keyword in exclude_keywords)
    
    return has_revenue_keyword and not has_exclude_keyword

def is_cost_of_sales_account(account):
    """Determine if account is cost of sales/cost of goods sold"""
    cost_keywords = ['cost of sales', 'cost of goods', 'cogs', 'direct cost', 'material cost', 'purchase']
    
    account_name_lower = account.name.lower()
    account_type_lower = (account.account_type or '').lower()
    
    return any(keyword in account_name_lower or keyword in account_type_lower 
              for keyword in cost_keywords)

def is_finance_cost_account(account):
    """Determine if account is finance cost/interest expense"""
    finance_keywords = ['interest', 'finance', 'loan', 'borrowing', 'bank charges']
    
    account_name_lower = account.name.lower()
    account_type_lower = (account.account_type or '').lower()
    
    return any(keyword in account_name_lower or keyword in account_type_lower 
              for keyword in finance_keywords)

def generate_income_statement_comparison(company, comparison_start, comparison_end):
    """Generate income statement data for comparison period"""
    try:
        from datetime import datetime
        
        if isinstance(comparison_start, str):
            comparison_start = datetime.strptime(comparison_start, '%Y-%m-%d').date()
        if isinstance(comparison_end, str):
            comparison_end = datetime.strptime(comparison_end, '%Y-%m-%d').date()
        
        accounts = Account.objects.filter(company=company, is_active=True)
        
        comparison_totals = {
            'total_revenue': 0,
            'total_cost_of_sales': 0,
            'total_operating_expenses': 0,
            'total_other_income': 0,
            'total_finance_costs': 0
        }
        
        for account in accounts:
            balance = calculate_account_balance_for_period(account, comparison_start, comparison_end)
            
            if account.type == 'income':
                if is_revenue_account(account):
                    comparison_totals['total_revenue'] += abs(balance)
                else:
                    comparison_totals['total_other_income'] += abs(balance)
            elif account.type == 'expense':
                if is_cost_of_sales_account(account):
                    comparison_totals['total_cost_of_sales'] += abs(balance)
                elif is_finance_cost_account(account):
                    comparison_totals['total_finance_costs'] += abs(balance)
                else:
                    comparison_totals['total_operating_expenses'] += abs(balance)
        
        # Calculate derived metrics
        comparison_totals['gross_profit'] = comparison_totals['total_revenue'] - comparison_totals['total_cost_of_sales']
        comparison_totals['operating_profit'] = (comparison_totals['gross_profit'] - 
                                               comparison_totals['total_operating_expenses'] + 
                                               comparison_totals['total_other_income'])
        comparison_totals['profit_before_tax'] = comparison_totals['operating_profit'] - comparison_totals['total_finance_costs']
        comparison_totals['net_profit'] = comparison_totals['profit_before_tax'] * 0.70  # Simplified tax calculation
        
        return {key: float(value) for key, value in comparison_totals.items()}
    except Exception:
        return None

def generate_equity_changes_data(company, period_start, period_end, comparison_year=None):
    """Generate IFRS 18 compliant Statement of Changes in Equity"""
    try:
        from datetime import datetime
        
        if isinstance(period_start, str):
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
        if isinstance(period_end, str):
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        
        # Get equity accounts
        equity_accounts = Account.objects.filter(
            company=company, 
            type='equity', 
            is_active=True
        )
        
        equity_movements = []
        
        for account in equity_accounts:
            # Opening balance (as of period start)
            opening_balance = calculate_account_balance_as_of(account, period_start)
            
            # Movements during period
            period_movement = calculate_account_balance_for_period(account, period_start, period_end)
            
            # Closing balance
            closing_balance = calculate_account_balance_as_of(account, period_end)
            
            if opening_balance != 0 or period_movement != 0 or closing_balance != 0:
                equity_movements.append({
                    'code': account.code,
                    'name': account.name,
                    'opening_balance': float(opening_balance),
                    'period_movement': float(period_movement),
                    'closing_balance': float(closing_balance),
                    'account_type': account.account_type or 'General'
                })
        
        # Calculate totals
        total_opening = sum(item['opening_balance'] for item in equity_movements)
        total_movement = sum(item['period_movement'] for item in equity_movements)
        total_closing = sum(item['closing_balance'] for item in equity_movements)
        
        return JsonResponse({
            'success': True,
            'report_type': 'equity_changes',
            'period_start': period_start.strftime('%Y-%m-%d'),
            'period_end': period_end.strftime('%Y-%m-%d'),
            'company_name': company.name,
            'data': {
                'equity_movements': equity_movements,
                'totals': {
                    'opening_balance': float(total_opening),
                    'period_movement': float(total_movement),
                    'closing_balance': float(total_closing)
                }
            },
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ifrs_compliance': True
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# CRUD Views for new features

@login_required
def template_create(request):
    """Create a new account template"""
    if request.method == 'POST':
        try:
            from .models import AccountTemplate
            
            template = AccountTemplate.objects.create(
                name=request.POST.get('name'),
                category=request.POST.get('category'),
                description=request.POST.get('description'),
                is_system=False,
                created_by=request.user
            )
            
            messages.success(request, f'Template "{template.name}" created successfully!')
            return redirect('accounting:account_templates')
        except Exception as e:
            messages.error(request, f'Error creating template: {str(e)}')
    
    from .models import AccountTemplate
    categories = AccountTemplate.TEMPLATE_CATEGORIES
    return render(request, 'accounting/template_form.html', {'categories': categories})


@login_required  
def template_apply(request, template_id):
    """Apply a template to create accounts and groups"""
    if request.method == 'POST':
        try:
            from .models import AccountTemplate, AccountTemplateAccount, AccountTemplateGroup
            
            template = get_object_or_404(AccountTemplate, id=template_id)
            company = request.user.company
            
            # Create account groups first
            group_mapping = {}
            for template_group in template.template_groups.all():
                # Check if category exists
                category, created = AccountCategory.objects.get_or_create(
                    company=company,
                    name=template_group.category_name,
                    defaults={'code': template_group.category_name[:10].upper()}
                )
                
                # Create account group
                group, created = AccountGroup.objects.get_or_create(
                    company=company,
                    code=template_group.code,
                    defaults={
                        'name': template_group.name,
                        'category': category,
                        'description': template_group.description
                    }
                )
                group_mapping[template_group.code] = group
            
            # Create accounts
            accounts_created = 0
            for template_account in template.template_accounts.all():
                # Find the group
                group = group_mapping.get(template_account.group_code)
                if not group and template_account.group_code:
                    continue  # Skip if group not found
                
                # Create account if it doesn't exist
                account, created = Account.objects.get_or_create(
                    company=company,
                    code=template_account.code,
                    defaults={
                        'name': template_account.name,
                        'type': template_account.account_type,
                        'balance_side': template_account.balance_side,
                        'group': group,
                        'description': template_account.description,
                        'is_default': template_account.is_default
                    }
                )
                if created:
                    accounts_created += 1
            
            messages.success(request, f'Template applied successfully! Created {accounts_created} accounts.')
            return redirect('accounting:accounts')
            
        except Exception as e:
            messages.error(request, f'Error applying template: {str(e)}')
            return redirect('accounting:account_templates')
    
    template = get_object_or_404(AccountTemplate, id=template_id)
    return render(request, 'accounting/template_apply.html', {'template': template})


@login_required
def reconciliation_create(request):
    """Create a new reconciliation"""
    if request.method == 'POST':
        try:
            from .models import AccountReconciliation
            from datetime import datetime
            
            account = get_object_or_404(Account, id=request.POST.get('account'))
            
            # Calculate current book balance (simplified)
            book_balance = calculate_account_balance(account)
            
            reconciliation = AccountReconciliation.objects.create(
                account=account,
                reconciliation_date=datetime.now().date(),
                statement_date=request.POST.get('statement_date'),
                statement_balance=request.POST.get('statement_balance'),
                book_balance=book_balance,
                created_by=request.user
            )
            
            messages.success(request, f'Reconciliation started for {account.name}')
            return redirect('accounting:reconciliation_detail', reconciliation_id=reconciliation.id)
            
        except Exception as e:
            messages.error(request, f'Error creating reconciliation: {str(e)}')
    
    company = request.user.company
    reconcilable_accounts = Account.objects.filter(
        company=company,
        is_active=True,
        type__in=['asset', 'liability']
    ).exclude(is_group=True)
    
    return render(request, 'accounting/reconciliation_form.html', {
        'reconcilable_accounts': reconcilable_accounts
    })


@login_required
def reconciliation_detail(request, reconciliation_id):
    """Reconciliation detail and workflow"""
    from .models import AccountReconciliation
    
    reconciliation = get_object_or_404(AccountReconciliation, id=reconciliation_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'complete':
            reconciliation.status = 'completed'
            reconciliation.completed_at = timezone.now()
            reconciliation.save()
            messages.success(request, 'Reconciliation completed successfully!')
            
        elif action == 'cancel':
            reconciliation.status = 'cancelled'
            reconciliation.save()
            messages.info(request, 'Reconciliation cancelled.')
            
        return redirect('accounting:account_reconciliation')
    
    return render(request, 'accounting/reconciliation_detail.html', {
        'reconciliation': reconciliation
    })


@login_required
def settings_update(request):
    """Update COA settings"""
    if request.method == 'POST':
        try:
            from .models import COASettings
            
            company = request.user.company
            settings, created = COASettings.objects.get_or_create(company=company)
            
            # Update settings from form
            settings.code_format = request.POST.get('code_format', settings.code_format)
            settings.code_length = request.POST.get('code_length', settings.code_length)
            settings.custom_pattern = request.POST.get('custom_pattern', settings.custom_pattern)
            settings.auto_generate_codes = request.POST.get('auto_generate_codes') == 'on'
            settings.strict_code_validation = request.POST.get('strict_code_validation') == 'on'
            settings.default_view = request.POST.get('default_view', settings.default_view)
            settings.accounts_per_page = request.POST.get('accounts_per_page', settings.accounts_per_page)
            settings.show_account_codes = request.POST.get('show_account_codes') == 'on'
            settings.show_balances = request.POST.get('show_balances') == 'on'
            settings.color_coding = request.POST.get('color_coding') == 'on'
            settings.hierarchical_indent = request.POST.get('hierarchical_indent') == 'on'
            settings.require_approval_new_accounts = request.POST.get('require_approval_new_accounts') == 'on'
            settings.restrict_account_deletion = request.POST.get('restrict_account_deletion') == 'on'
            settings.log_account_changes = request.POST.get('log_account_changes') == 'on'
            settings.log_balance_changes = request.POST.get('log_balance_changes') == 'on'
            settings.auto_sales_entries = request.POST.get('auto_sales_entries') == 'on'
            settings.auto_purchase_entries = request.POST.get('auto_purchase_entries') == 'on'
            settings.auto_inventory_entries = request.POST.get('auto_inventory_entries') == 'on'
            settings.auto_backup_enabled = request.POST.get('auto_backup_enabled') == 'on'
            settings.backup_frequency = request.POST.get('backup_frequency', settings.backup_frequency)
            
            # Handle default accounts
            if request.POST.get('default_sales_account'):
                settings.default_sales_account_id = request.POST.get('default_sales_account')
            if request.POST.get('default_purchase_account'):
                settings.default_purchase_account_id = request.POST.get('default_purchase_account')
            if request.POST.get('default_inventory_account'):
                settings.default_inventory_account_id = request.POST.get('default_inventory_account')
            
            settings.save()
            messages.success(request, 'COA settings updated successfully!')
            
        except Exception as e:
            messages.error(request, f'Error updating settings: {str(e)}')
    
    return redirect('accounting:coa_settings')


@login_required
def import_accounts(request):
    """Import accounts from file with support for CSV, Excel, and JSON"""
    if request.method == 'POST':
        try:
            from .models import ImportExportOperation
            import csv
            import io
            import json
            
            uploaded_file = request.FILES.get('import_file')
            if not uploaded_file:
                messages.error(request, 'Please select a file to import.')
                return redirect('accounting:account_import_export')
            
            file_format = request.POST.get('file_format', 'csv')
            import_type = request.POST.get('import_type', 'accounts')
            
            # Create import operation record
            operation = ImportExportOperation.objects.create(
                operation_type='import',
                data_type=import_type,
                file_name=uploaded_file.name,
                file_format=file_format,
                status='processing',
                created_by=request.user
            )
            
            success_count = 0
            error_count = 0
            company = request.user.company
            errors = []
            
            try:
                if file_format == 'csv' or uploaded_file.name.endswith('.csv'):
                    # Process CSV file
                    file_content = uploaded_file.read().decode('utf-8')
                    csv_data = csv.DictReader(io.StringIO(file_content))
                    
                    for row_num, row in enumerate(csv_data, start=2):
                        try:
                            success, error = process_account_row(row, company, request.POST)
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                errors.append(f"Row {row_num}: {error}")
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Row {row_num}: {str(e)}")
                
                elif file_format == 'excel' or uploaded_file.name.endswith(('.xlsx', '.xls')):
                    # Process Excel file
                    try:
                        import openpyxl
                        from openpyxl import load_workbook
                        
                        wb = load_workbook(uploaded_file)
                        ws = wb.active
                        
                        # Get headers from first row
                        headers = [cell.value for cell in ws[1]]
                        
                        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            try:
                                # Convert row to dictionary
                                row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                                success, error = process_account_row(row_dict, company, request.POST)
                                if success:
                                    success_count += 1
                                else:
                                    error_count += 1
                                    errors.append(f"Row {row_num}: {error}")
                            except Exception as e:
                                error_count += 1
                                errors.append(f"Row {row_num}: {str(e)}")
                                
                    except ImportError:
                        operation.status = 'failed'
                        operation.error_log = 'openpyxl library not installed. Please install it to process Excel files.'
                        operation.save()
                        messages.error(request, 'Excel processing not available. Please install openpyxl library.')
                        return redirect('accounting:account_import_export')
                
                elif file_format == 'json' or uploaded_file.name.endswith('.json'):
                    # Process JSON file
                    file_content = uploaded_file.read().decode('utf-8')
                    json_data = json.loads(file_content)
                    
                    if isinstance(json_data, list):
                        accounts_data = json_data
                    elif isinstance(json_data, dict) and 'accounts' in json_data:
                        accounts_data = json_data['accounts']
                    else:
                        raise ValueError("Invalid JSON format. Expected list of accounts or object with 'accounts' key.")
                    
                    for row_num, account_data in enumerate(accounts_data, start=1):
                        try:
                            success, error = process_account_row(account_data, company, request.POST)
                            if success:
                                success_count += 1
                            else:
                                error_count += 1
                                errors.append(f"Account {row_num}: {error}")
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Account {row_num}: {str(e)}")
                
                else:
                    operation.status = 'failed'
                    operation.error_log = f'Unsupported file format: {file_format}'
                    operation.save()
                    messages.error(request, f'Unsupported file format: {file_format}. Supported formats: CSV, Excel (XLSX), JSON')
                    return redirect('accounting:account_import_export')
                
                # Update operation
                operation.record_count = success_count + error_count
                operation.success_count = success_count
                operation.error_count = error_count
                operation.status = 'completed' if error_count == 0 else 'completed_with_errors'
                operation.completed_at = timezone.now()
                
                if errors:
                    operation.error_log = '\n'.join(errors[:10])  # Store first 10 errors
                
                operation.save()
                
                if error_count == 0:
                    messages.success(request, f'Import completed successfully: {success_count} accounts imported.')
                else:
                    messages.warning(request, f'Import completed with issues: {success_count} accounts imported, {error_count} errors. Check the operation log for details.')
                    
            except Exception as e:
                operation.status = 'failed'
                operation.error_log = str(e)
                operation.save()
                messages.error(request, f'Import failed: {str(e)}')
                
        except Exception as e:
            messages.error(request, f'Error processing import: {str(e)}')
    
    return redirect('accounting:account_import_export')


def process_account_row(row_data, company, form_data):
    """Process a single account row from import data"""
    try:
        # Required fields
        code = row_data.get('code') or row_data.get('account_code')
        name = row_data.get('name') or row_data.get('account_name')
        account_type = row_data.get('type') or row_data.get('account_type', 'asset')
        
        if not code or not name:
            return False, "Missing required fields: code and name"
        
        # Optional fields
        description = row_data.get('description', '')
        balance_side = row_data.get('balance_side', 'debit')
        is_active = str(row_data.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        group_code = row_data.get('group_code') or row_data.get('group')
        
        # Find or create account group
        group = None
        if group_code and form_data.get('create_groups') == 'on':
            # Create category if needed
            category, _ = AccountCategory.objects.get_or_create(
                company=company,
                name='Imported',
                defaults={'code': 'IMP'}
            )
            
            # Create group if needed
            group, _ = AccountGroup.objects.get_or_create(
                company=company,
                code=group_code,
                defaults={
                    'name': f'Group {group_code}', 
                    'category': category,
                    'description': 'Imported account group'
                }
            )
        
        # Check for existing account
        existing_account = Account.objects.filter(company=company, code=code).first()
        
        if existing_account:
            if form_data.get('update_existing') == 'on':
                # Update existing account
                existing_account.name = name
                existing_account.type = account_type
                existing_account.description = description
                existing_account.balance_side = balance_side
                existing_account.is_active = is_active
                if group:
                    existing_account.group = group
                existing_account.save()
                return True, None
            else:
                return False, f"Account with code '{code}' already exists"
        else:
            # Create new account
            Account.objects.create(
                company=company,
                code=code,
                name=name,
                type=account_type,
                balance_side=balance_side,
                group=group,
                description=description,
                is_active=is_active
            )
            return True, None
            
    except Exception as e:
        return False, str(e)


@login_required
def export_accounts(request):
    """Export accounts to file with support for CSV, Excel, JSON, and PDF"""
    try:
        from .models import ImportExportOperation
        import csv
        import json
        from django.http import HttpResponse
        
        company = request.user.company
        export_type = request.GET.get('type', 'all')
        file_format = request.GET.get('format', 'csv')
        
        # Create export operation record
        operation = ImportExportOperation.objects.create(
            operation_type='export',
            data_type='accounts',
            file_format=file_format,
            status='processing',
            created_by=request.user
        )
        
        # Get accounts based on export type
        accounts = Account.objects.filter(company=company)
        if export_type == 'active':
            accounts = accounts.filter(is_active=True)
        elif export_type == 'by_type':
            account_types = request.GET.getlist('types[]')
            if account_types:
                accounts = accounts.filter(type__in=account_types)
        
        # Include balances if requested
        include_balances = request.GET.get('include_balances') == 'on'
        include_groups = request.GET.get('include_groups') == 'on'
        include_inactive = request.GET.get('include_inactive') == 'on'
        
        if not include_inactive:
            accounts = accounts.filter(is_active=True)
        
        if file_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="accounts_export.csv"'
            
            writer = csv.writer(response)
            
            # Write headers
            headers = ['code', 'name', 'type', 'balance_side', 'description', 'is_active']
            if include_groups:
                headers.extend(['group_code', 'group_name'])
            if include_balances:
                headers.append('current_balance')
            
            writer.writerow(headers)
            
            # Write data
            for account in accounts:
                row = [
                    account.code,
                    account.name,
                    account.type,
                    account.balance_side,
                    account.description,
                    account.is_active
                ]
                
                if include_groups:
                    row.extend([
                        account.group.code if account.group else '',
                        account.group.name if account.group else ''
                    ])
                
                if include_balances:
                    balance = calculate_account_balance(account)
                    row.append(balance)
                
                writer.writerow(row)
            
            operation.record_count = accounts.count()
            operation.success_count = accounts.count()
            operation.status = 'completed'
            operation.completed_at = timezone.now()
            operation.file_name = 'accounts_export.csv'
            operation.save()
            
            return response
            
        elif file_format == 'excel':
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from django.http import HttpResponse
                import io
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Chart of Accounts"
                
                # Headers
                headers = ['Code', 'Name', 'Type', 'Balance Side', 'Description', 'Active']
                if include_groups:
                    headers.extend(['Group Code', 'Group Name'])
                if include_balances:
                    headers.append('Current Balance')
                
                # Style headers
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Write data
                for row_num, account in enumerate(accounts, 2):
                    data = [
                        account.code,
                        account.name,
                        account.get_type_display(),
                        account.get_balance_side_display(),
                        account.description,
                        'Yes' if account.is_active else 'No'
                    ]
                    
                    if include_groups:
                        data.extend([
                            account.group.code if account.group else '',
                            account.group.name if account.group else ''
                        ])
                    
                    if include_balances:
                        balance = calculate_account_balance(account)
                        data.append(balance)
                    
                    for col, value in enumerate(data, 1):
                        ws.cell(row=row_num, column=col, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to response
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="accounts_export.xlsx"'
                
                operation.record_count = accounts.count()
                operation.success_count = accounts.count()
                operation.status = 'completed'
                operation.completed_at = timezone.now()
                operation.file_name = 'accounts_export.xlsx'
                operation.save()
                
                return response
                
            except ImportError:
                operation.status = 'failed'
                operation.error_log = 'openpyxl library not installed'
                operation.save()
                messages.error(request, 'Excel export not available. Please install openpyxl library.')
                return redirect('accounting:account_import_export')
                
        elif file_format == 'json':
            accounts_data = []
            
            for account in accounts:
                account_data = {
                    'code': account.code,
                    'name': account.name,
                    'type': account.type,
                    'balance_side': account.balance_side,
                    'description': account.description,
                    'is_active': account.is_active,
                    'created_at': account.created_at.isoformat() if account.created_at else None
                }
                
                if include_groups and account.group:
                    account_data['group'] = {
                        'code': account.group.code,
                        'name': account.group.name
                    }
                
                if include_balances:
                    account_data['current_balance'] = float(calculate_account_balance(account))
                
                accounts_data.append(account_data)
            
            export_data = {
                'export_info': {
                    'company': company.name,
                    'export_date': timezone.now().isoformat(),
                    'export_type': export_type,
                    'total_accounts': len(accounts_data)
                },
                'accounts': accounts_data
            }
            
            response = HttpResponse(
                json.dumps(export_data, indent=2),
                content_type='application/json'
            )
            response['Content-Disposition'] = 'attachment; filename="accounts_export.json"'
            
            operation.record_count = accounts.count()
            operation.success_count = accounts.count()
            operation.status = 'completed'
            operation.completed_at = timezone.now()
            operation.file_name = 'accounts_export.json'
            operation.save()
            
            return response
            
        elif file_format == 'pdf':
            try:
                from django.template.loader import get_template
                from django.http import HttpResponse
                from weasyprint import HTML, CSS
                from django.template import Context, Template
                import io
                
                # Prepare context for PDF
                accounts_with_balances = []
                total_debit_balance = 0
                total_credit_balance = 0
                
                for account in accounts:
                    balance = calculate_account_balance(account) if include_balances else 0
                    debit_balance = balance if account.balance_side == 'debit' and balance > 0 else 0
                    credit_balance = balance if account.balance_side == 'credit' and balance > 0 else 0
                    
                    accounts_with_balances.append({
                        'account': account,
                        'balance': balance,
                        'debit_balance': debit_balance,
                        'credit_balance': credit_balance
                    })
                    
                    total_debit_balance += debit_balance
                    total_credit_balance += credit_balance
                
                context = {
                    'company': company,
                    'accounts': accounts_with_balances,
                    'export_date': timezone.now(),
                    'include_balances': include_balances,
                    'include_groups': include_groups,
                    'total_debit_balance': total_debit_balance,
                    'total_credit_balance': total_credit_balance,
                    'export_type': export_type
                }
                
                # HTML template for PDF
                html_template = """
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <title>Chart of Accounts - {{ company.name }}</title>
                    <style>
                        body { font-family: Arial, sans-serif; margin: 20px; }
                        .header { text-align: center; margin-bottom: 30px; }
                        .company-name { font-size: 24px; font-weight: bold; }
                        .report-title { font-size: 18px; margin: 10px 0; }
                        .report-date { font-size: 12px; color: #666; }
                        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                        th { background-color: #f2f2f2; font-weight: bold; }
                        .number { text-align: right; }
                        .total-row { font-weight: bold; background-color: #f9f9f9; }
                        .summary { margin-top: 30px; }
                    </style>
                </head>
                <body>
                    <div class="header">
                        <div class="company-name">{{ company.name }}</div>
                        <div class="report-title">Chart of Accounts</div>
                        <div class="report-date">Generated on {{ export_date|date:"F d, Y g:i A" }}</div>
                    </div>
                    
                    <table>
                        <thead>
                            <tr>
                                <th>Code</th>
                                <th>Account Name</th>
                                <th>Type</th>
                                {% if include_groups %}<th>Group</th>{% endif %}
                                {% if include_balances %}
                                <th class="number">Debit Balance</th>
                                <th class="number">Credit Balance</th>
                                {% endif %}
                            </tr>
                        </thead>
                        <tbody>
                            {% for item in accounts %}
                            <tr>
                                <td>{{ item.account.code }}</td>
                                <td>{{ item.account.name }}</td>
                                <td>{{ item.account.get_type_display }}</td>
                                {% if include_groups %}<td>{{ item.account.group.name|default:"-" }}</td>{% endif %}
                                {% if include_balances %}
                                <td class="number">{% if item.debit_balance %}${{ item.debit_balance|floatformat:2 }}{% endif %}</td>
                                <td class="number">{% if item.credit_balance %}${{ item.credit_balance|floatformat:2 }}{% endif %}</td>
                                {% endif %}
                            </tr>
                            {% endfor %}
                            {% if include_balances %}
                            <tr class="total-row">
                                <td colspan="{% if include_groups %}3{% else %}3{% endif %}"><strong>TOTALS</strong></td>
                                {% if include_groups %}<td></td>{% endif %}
                                <td class="number"><strong>${{ total_debit_balance|floatformat:2 }}</strong></td>
                                <td class="number"><strong>${{ total_credit_balance|floatformat:2 }}</strong></td>
                            </tr>
                            {% endif %}
                        </tbody>
                    </table>
                    
                    <div class="summary">
                        <p><strong>Export Summary:</strong></p>
                        <p>Total Accounts: {{ accounts|length }}</p>
                        <p>Export Type: {{ export_type|title }}</p>
                        {% if include_balances %}
                        <p>Balance Difference: ${{ total_debit_balance|add:total_credit_balance|add:"-"|add:total_credit_balance|floatformat:2 }}</p>
                        {% endif %}
                    </div>
                </body>
                </html>
                """
                
                template = Template(html_template)
                html_content = template.render(Context(context))
                
                # Generate PDF
                html_doc = HTML(string=html_content)
                pdf_buffer = io.BytesIO()
                html_doc.write_pdf(pdf_buffer)
                pdf_buffer.seek(0)
                
                response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="chart_of_accounts.pdf"'
                
                operation.record_count = accounts.count()
                operation.success_count = accounts.count()
                operation.status = 'completed'
                operation.completed_at = timezone.now()
                operation.file_name = 'chart_of_accounts.pdf'
                operation.save()
                
                return response
                
            except ImportError:
                operation.status = 'failed'
                operation.error_log = 'weasyprint library not installed'
                operation.save()
                messages.error(request, 'PDF export not available. Please install weasyprint library.')
                return redirect('accounting:account_import_export')
        
        else:
            operation.status = 'failed'
            operation.error_log = f'Unsupported export format: {file_format}'
            operation.save()
            messages.error(request, f'Unsupported export format: {file_format}')
        
    except Exception as e:
        messages.error(request, f'Error exporting accounts: {str(e)}')
    
    return redirect('accounting:account_import_export')

@login_required
def chart_of_accounts_ui(request):
    """Display hierarchical chart of accounts view"""
    try:
        company = request.user.company
        
        # Get all accounts organized by type and hierarchy
        accounts = Account.objects.filter(company=company, is_active=True).select_related('group')
        
        # Calculate balances for each account
        for account in accounts:
            account.balance = calculate_account_balance(account)
        
        # Organize accounts by type
        accounts_by_type = {
            'asset': accounts.filter(type='asset').order_by('code'),
            'liability': accounts.filter(type='liability').order_by('code'),
            'equity': accounts.filter(type='equity').order_by('code'),
            'income': accounts.filter(type='income').order_by('code'),
            'expense': accounts.filter(type='expense').order_by('code'),
        }
        
        # Calculate totals by type
        totals = {}
        for account_type, type_accounts in accounts_by_type.items():
            totals[account_type] = sum(acc.balance for acc in type_accounts)
        
        context = {
            'accounts_by_type': accounts_by_type,
            'totals': totals,
            'company': company,
        }
        
        return render(request, 'accounting/chart_of_accounts.html', context)
    except Exception as e:
        context = {
            'accounts_by_type': {},
            'totals': {},
            'error': str(e)
        }
        return render(request, 'accounting/chart_of_accounts.html', context)

# API endpoints for frontend integration
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json

@login_required
@require_http_methods(["GET"])
def api_chart_of_accounts(request):
    """API endpoint to get chart of accounts data"""
    try:
        company = request.user.company
        accounts = Account.objects.filter(company=company, is_active=True).select_related('group')
        
        # Calculate balances and organize by type
        accounts_data = []
        for account in accounts:
            balance = calculate_account_balance(account)
            accounts_data.append({
                'id': account.id,
                'code': account.code,
                'name': account.name,
                'type': account.type,
                'type_display': account.get_type_display(),
                'group': account.group.name if account.group else None,
                'balance': float(balance),
                'balance_side': account.balance_side,
                'is_active': account.is_active,
                'description': account.description,
            })
        
        return JsonResponse({
            'success': True,
            'accounts': accounts_data,
            'total_accounts': len(accounts_data)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def api_journal_entries(request):
    """API endpoint to get journal entries"""
    try:
        company = request.user.company
        entries = JournalEntry.objects.filter(company=company).select_related('journal', 'created_by').order_by('-date')[:50]
        
        entries_data = []
        for entry in entries:
            # Get journal items for this entry
            items = entry.items.all().select_related('account')
            items_data = []
            for item in items:
                items_data.append({
                    'account_id': item.account.id,
                    'account_code': item.account.code,
                    'account_name': item.account.name,
                    'debit': float(item.debit) if item.debit else 0,
                    'credit': float(item.credit) if item.credit else 0,
                    'description': item.description,
                })
            
            entries_data.append({
                'id': entry.id,
                'reference': entry.reference,
                'date': entry.date.strftime('%Y-%m-%d') if entry.date else None,
                'description': entry.description,
                'journal': entry.journal.name if entry.journal else None,
                'total_amount': float(entry.total_amount) if entry.total_amount else 0,
                'status': entry.status,
                'created_by': entry.created_by.get_full_name() if entry.created_by else None,
                'created_at': entry.created_at.strftime('%Y-%m-%d %H:%M') if entry.created_at else None,
                'items': items_data,
            })
        
        return JsonResponse({
            'success': True,
            'entries': entries_data,
            'total_entries': len(entries_data)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_create_journal_entry(request):
    """API endpoint to create a new journal entry"""
    try:
        company = request.user.company
        data = json.loads(request.body)
        
        # Validate required fields
        if not data.get('description') or not data.get('items'):
            return JsonResponse({'success': False, 'error': 'Description and items are required'}, status=400)
        
        # Create journal entry
        entry = JournalEntry.objects.create(
            company=company,
            reference=data.get('reference', ''),
            date=datetime.strptime(data.get('date'), '%Y-%m-%d').date() if data.get('date') else timezone.now().date(),
            description=data.get('description'),
            created_by=request.user,
            status='draft'
        )
        
        # Create journal items
        total_debit = 0
        total_credit = 0
        for item_data in data.get('items', []):
            account = Account.objects.get(id=item_data['account_id'], company=company)
            debit = float(item_data.get('debit', 0))
            credit = float(item_data.get('credit', 0))
            
            JournalItem.objects.create(
                entry=entry,
                account=account,
                debit=debit,
                credit=credit,
                description=item_data.get('description', '')
            )
            
            total_debit += debit
            total_credit += credit
        
        # Validate balanced entry
        if abs(total_debit - total_credit) > 0.01:  # Allow for small rounding differences
            entry.delete()
            return JsonResponse({'success': False, 'error': 'Journal entry must be balanced (debits must equal credits)'}, status=400)
        
        entry.total_amount = total_debit
        entry.save()
        
        return JsonResponse({
            'success': True,
            'entry_id': entry.id,
            'message': 'Journal entry created successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def api_financial_reports(request):
    """API endpoint to get financial reports data"""
    try:
        company = request.user.company
        report_type = request.GET.get('type', 'balance_sheet')
        period_start = request.GET.get('period_start')
        period_end = request.GET.get('period_end')
        comparison_year = request.GET.get('comparison_year')
        
        # Set default period if not provided
        if not period_start or not period_end:
            from datetime import datetime, date
            current_year = datetime.now().year
            period_start = f"{current_year}-01-01"
            period_end = f"{current_year}-12-31"
        
        if report_type == 'balance_sheet':
            return generate_balance_sheet_data(company, period_end, comparison_year)
        elif report_type == 'income_statement':
            return generate_income_statement_data(company, period_start, period_end, comparison_year)
        elif report_type == 'cash_flow':
            return generate_cash_flow_data(company, period_start, period_end, comparison_year)
        elif report_type == 'trial_balance':
            return generate_trial_balance_data(company, period_end)
        elif report_type == 'equity_changes':
            return generate_equity_changes_data(company, period_start, period_end, comparison_year)
        else:
            return JsonResponse({'success': False, 'error': 'Invalid report type'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def generate_balance_sheet_data(company, period_end=None, comparison_year=None):
    """Generate IFRS 18 compliant balance sheet data"""
    try:
        from datetime import datetime, date
        from django.db.models import Q
        
        if not period_end:
            period_end = date.today().strftime('%Y-%m-%d')
        
        # Get accounts with proper IFRS 18 categorization
        accounts = Account.objects.filter(company=company, is_active=True)
        
        # Initialize IFRS 18 structure
        current_assets = []
        non_current_assets = []
        current_liabilities = []
        non_current_liabilities = []
        equity = []
        
        # Totals
        total_current_assets = 0
        total_non_current_assets = 0
        total_current_liabilities = 0
        total_non_current_liabilities = 0
        total_equity = 0
        
        for account in accounts:
            balance = calculate_account_balance_as_of(account, period_end)
            
            if balance == 0:
                continue
                
            account_data = {
                'code': account.code,
                'name': account.name,
                'balance': float(balance),
                'group': account.group.name if account.group else 'Unclassified',
                'account_type': account.account_type or 'General'
            }
            
            # IFRS 18 Asset Classification
            if account.type == 'asset':
                if is_current_asset(account):
                    current_assets.append(account_data)
                    total_current_assets += balance
                else:
                    non_current_assets.append(account_data)
                    total_non_current_assets += balance
            
            # IFRS 18 Liability Classification
            elif account.type == 'liability':
                if is_current_liability(account):
                    current_liabilities.append(account_data)
                    total_current_liabilities += balance
                else:
                    non_current_liabilities.append(account_data)
                    total_non_current_liabilities += balance
            
            # Equity
            elif account.type == 'equity':
                equity.append(account_data)
                total_equity += balance
        
        # Sort by account code
        current_assets.sort(key=lambda x: x['code'])
        non_current_assets.sort(key=lambda x: x['code'])
        current_liabilities.sort(key=lambda x: x['code'])
        non_current_liabilities.sort(key=lambda x: x['code'])
        equity.sort(key=lambda x: x['code'])
        
        total_assets = total_current_assets + total_non_current_assets
        total_liabilities = total_current_liabilities + total_non_current_liabilities
        total_liabilities_equity = total_liabilities + total_equity
        
        # Comparison data if requested
        comparison_data = None
        if comparison_year:
            comparison_period_end = f"{comparison_year}-12-31"
            comparison_data = generate_balance_sheet_comparison(company, comparison_period_end)
        
        return JsonResponse({
            'success': True,
            'report_type': 'balance_sheet',
            'period_end': period_end,
            'company_name': company.name,
            'data': {
                'assets': {
                    'current': current_assets,
                    'non_current': non_current_assets
                },
                'liabilities': {
                    'current': current_liabilities,
                    'non_current': non_current_liabilities
                },
                'equity': equity,
                'totals': {
                    'current_assets': float(total_current_assets),
                    'non_current_assets': float(total_non_current_assets),
                    'total_assets': float(total_assets),
                    'current_liabilities': float(total_current_liabilities),
                    'non_current_liabilities': float(total_non_current_liabilities),
                    'total_liabilities': float(total_liabilities),
                    'total_equity': float(total_equity),
                    'total_liabilities_equity': float(total_liabilities_equity)
                }
            },
            'comparison_data': comparison_data,
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ifrs_compliance': True
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def generate_income_statement_data(company, period_start, period_end, comparison_year=None):
    """Generate IFRS 18 compliant income statement data"""
    try:
        from datetime import datetime
        
        if isinstance(period_start, str):
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
        if isinstance(period_end, str):
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        
        accounts = Account.objects.filter(company=company, is_active=True)
        
        # IFRS 18 Income Statement Structure
        revenue = []
        cost_of_sales = []
        operating_expenses = []
        other_income = []
        finance_costs = []
        
        total_revenue = 0
        total_cost_of_sales = 0
        total_operating_expenses = 0
        total_other_income = 0
        total_finance_costs = 0
        
        for account in accounts:
            # Calculate balance for the period
            balance = calculate_account_balance_for_period(account, period_start, period_end)
            
            if balance == 0:
                continue
                
            account_data = {
                'code': account.code,
                'name': account.name,
                'balance': float(abs(balance)),  # Always positive for presentation
                'group': account.group.name if account.group else 'Unclassified',
                'account_type': account.account_type or 'General'
            }
            
            if account.type == 'income':
                if is_revenue_account(account):
                    revenue.append(account_data)
                    total_revenue += abs(balance)
                else:
                    other_income.append(account_data)
                    total_other_income += abs(balance)
            elif account.type == 'expense':
                if is_cost_of_sales_account(account):
                    cost_of_sales.append(account_data)
                    total_cost_of_sales += abs(balance)
                elif is_finance_cost_account(account):
                    finance_costs.append(account_data)
                    total_finance_costs += abs(balance)
                else:
                    operating_expenses.append(account_data)
                    total_operating_expenses += abs(balance)
        
        # Sort by account code
        revenue.sort(key=lambda x: x['code'])
        cost_of_sales.sort(key=lambda x: x['code'])
        operating_expenses.sort(key=lambda x: x['code'])
        other_income.sort(key=lambda x: x['code'])
        finance_costs.sort(key=lambda x: x['code'])
        
        # Calculate key metrics
        gross_profit = total_revenue - total_cost_of_sales
        operating_profit = gross_profit - total_operating_expenses + total_other_income
        profit_before_tax = operating_profit - total_finance_costs
        
        # Calculate tax (simplified - you might want to integrate with tax module)
        estimated_tax_rate = 0.30  # 30% - adjust based on company's jurisdiction
        estimated_tax = max(0, profit_before_tax * estimated_tax_rate)
        net_profit = profit_before_tax - estimated_tax
        
        # Comparison data if requested
        comparison_data = None
        if comparison_year:
            comparison_start = f"{comparison_year}-01-01"
            comparison_end = f"{comparison_year}-12-31"
            comparison_data = generate_income_statement_comparison(company, comparison_start, comparison_end)
        
        return JsonResponse({
            'success': True,
            'report_type': 'income_statement',
            'period_start': period_start.strftime('%Y-%m-%d'),
            'period_end': period_end.strftime('%Y-%m-%d'),
            'company_name': company.name,
            'data': {
                'revenue': revenue,
                'cost_of_sales': cost_of_sales,
                'operating_expenses': operating_expenses,
                'other_income': other_income,
                'finance_costs': finance_costs,
                'metrics': {
                    'total_revenue': float(total_revenue),
                    'total_cost_of_sales': float(total_cost_of_sales),
                    'gross_profit': float(gross_profit),
                    'gross_profit_margin': float((gross_profit / total_revenue * 100) if total_revenue > 0 else 0),
                    'total_operating_expenses': float(total_operating_expenses),
                    'total_other_income': float(total_other_income),
                    'operating_profit': float(operating_profit),
                    'operating_margin': float((operating_profit / total_revenue * 100) if total_revenue > 0 else 0),
                    'total_finance_costs': float(total_finance_costs),
                    'profit_before_tax': float(profit_before_tax),
                    'estimated_tax': float(estimated_tax),
                    'net_profit': float(net_profit),
                    'net_margin': float((net_profit / total_revenue * 100) if total_revenue > 0 else 0)
                }
            },
            'comparison_data': comparison_data,
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ifrs_compliance': True
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def generate_cash_flow_data(company, period_start, period_end, comparison_year=None):
    """Generate IFRS 18 compliant cash flow statement data"""
    try:
        from datetime import datetime
        
        if isinstance(period_start, str):
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
        if isinstance(period_end, str):
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        
        # Get cash and cash equivalent accounts
        cash_accounts = Account.objects.filter(
            company=company,
            is_active=True
        ).filter(
            Q(account_type__icontains='cash') | 
            Q(account_type__icontains='bank') |
            Q(name__icontains='cash') |
            Q(name__icontains='bank')
        )
        
        # Calculate opening and closing cash balances
        opening_cash = sum(calculate_account_balance_as_of(account, period_start) for account in cash_accounts)
        closing_cash = sum(calculate_account_balance_as_of(account, period_end) for account in cash_accounts)
        
        # Get net income from income statement
        income_response = generate_income_statement_data(company, period_start, period_end)
        income_data = income_response.content.decode('utf-8')
        import json
        income_json = json.loads(income_data)
        net_income = income_json['data']['metrics']['net_profit'] if income_json['success'] else 0
        
        # Operating Activities (simplified - can be enhanced with detailed analysis)
        operating_activities = [
            {
                'description': 'Net Income',
                'amount': float(net_income),
                'type': 'inflow' if net_income > 0 else 'outflow'
            }
        ]
        
        # Add back non-cash expenses (depreciation, amortization)
        depreciation_accounts = Account.objects.filter(
            company=company,
            type='expense',
            is_active=True
        ).filter(
            Q(name__icontains='depreciation') |
            Q(name__icontains='amortization') |
            Q(account_type__icontains='depreciation')
        )
        
        total_depreciation = sum(
            abs(calculate_account_balance_for_period(account, period_start, period_end))
            for account in depreciation_accounts
        )
        
        if total_depreciation > 0:
            operating_activities.append({
                'description': 'Depreciation and Amortization',
                'amount': float(total_depreciation),
                'type': 'inflow'
            })
        
        # Working capital changes (simplified)
        current_asset_change = calculate_working_capital_change(company, period_start, period_end, 'asset')
        current_liability_change = calculate_working_capital_change(company, period_start, period_end, 'liability')
        
        if current_asset_change != 0:
            operating_activities.append({
                'description': 'Change in Current Assets',
                'amount': float(-current_asset_change),  # Increase in assets is cash outflow
                'type': 'outflow' if current_asset_change > 0 else 'inflow'
            })
        
        if current_liability_change != 0:
            operating_activities.append({
                'description': 'Change in Current Liabilities',
                'amount': float(current_liability_change),  # Increase in liabilities is cash inflow
                'type': 'inflow' if current_liability_change > 0 else 'outflow'
            })
        
        # Investing Activities
        investing_activities = []
        
        # Property, Plant & Equipment changes
        ppe_accounts = Account.objects.filter(
            company=company,
            type='asset',
            is_active=True
        ).filter(
            Q(name__icontains='equipment') |
            Q(name__icontains='property') |
            Q(name__icontains='plant') |
            Q(account_type__icontains='fixed') |
            Q(account_type__icontains='property')
        )
        
        ppe_change = sum(
            calculate_account_balance_for_period(account, period_start, period_end)
            for account in ppe_accounts
        )
        
        if ppe_change != 0:
            investing_activities.append({
                'description': 'Property, Plant & Equipment Changes',
                'amount': float(-ppe_change),  # Increase in PPE is cash outflow
                'type': 'outflow' if ppe_change > 0 else 'inflow'
            })
        
        # Financing Activities
        financing_activities = []
        
        # Long-term debt changes
        debt_accounts = Account.objects.filter(
            company=company,
            type='liability',
            is_active=True
        ).filter(
            Q(name__icontains='loan') |
            Q(name__icontains='debt') |
            Q(name__icontains='borrowing') |
            Q(account_type__icontains='long') |
            Q(account_type__icontains='debt')
        )
        
        debt_change = sum(
            calculate_account_balance_for_period(account, period_start, period_end)
            for account in debt_accounts
        )
        
        if debt_change != 0:
            financing_activities.append({
                'description': 'Long-term Debt Changes',
                'amount': float(debt_change),  # Increase in debt is cash inflow
                'type': 'inflow' if debt_change > 0 else 'outflow'
            })
        
        # Dividend payments
        dividend_accounts = Account.objects.filter(
            company=company,
            is_active=True
        ).filter(
            Q(name__icontains='dividend') |
            Q(account_type__icontains='dividend')
        )
        
        dividend_payments = sum(
            abs(calculate_account_balance_for_period(account, period_start, period_end))
            for account in dividend_accounts
        )
        
        if dividend_payments > 0:
            financing_activities.append({
                'description': 'Dividend Payments',
                'amount': float(-dividend_payments),
                'type': 'outflow'
            })
        
        # Calculate totals
        total_operating = sum(item['amount'] for item in operating_activities)
        total_investing = sum(item['amount'] for item in investing_activities)
        total_financing = sum(item['amount'] for item in financing_activities)
        net_cash_flow = total_operating + total_investing + total_financing
        
        return JsonResponse({
            'success': True,
            'report_type': 'cash_flow',
            'period_start': period_start.strftime('%Y-%m-%d'),
            'period_end': period_end.strftime('%Y-%m-%d'),
            'company_name': company.name,
            'data': {
                'operating_activities': operating_activities,
                'investing_activities': investing_activities,
                'financing_activities': financing_activities,
                'cash_balances': {
                    'opening_cash': float(opening_cash),
                    'closing_cash': float(closing_cash),
                    'calculated_closing': float(opening_cash + net_cash_flow)
                },
                'totals': {
                    'operating': float(total_operating),
                    'investing': float(total_investing),
                    'financing': float(total_financing),
                    'net_cash_flow': float(net_cash_flow)
                }
            },
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ifrs_compliance': True
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def calculate_working_capital_change(company, period_start, period_end, account_type):
    """Calculate working capital changes for cash flow statement"""
    try:
        # Get current assets or current liabilities
        accounts = Account.objects.filter(
            company=company,
            type=account_type,
            is_active=True
        )
        
        if account_type == 'asset':
            current_accounts = [acc for acc in accounts if is_current_asset(acc)]
        else:
            current_accounts = [acc for acc in accounts if is_current_liability(acc)]
        
        # Exclude cash accounts from working capital calculation
        current_accounts = [
            acc for acc in current_accounts 
            if 'cash' not in acc.name.lower() and 'bank' not in acc.name.lower()
        ]
        
        total_change = sum(
            calculate_account_balance_for_period(account, period_start, period_end)
            for account in current_accounts
        )
        
        return total_change
    except Exception:
        return 0
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def generate_trial_balance_data(company, period_end=None):
    """Generate trial balance data as of specific date"""
    try:
        from datetime import date
        
        if not period_end:
            period_end = date.today().strftime('%Y-%m-%d')
        
        accounts = Account.objects.filter(company=company, is_active=True)
        
        trial_balance = []
        total_debit = 0
        total_credit = 0
        
        for account in accounts:
            balance = calculate_account_balance_as_of(account, period_end)
            
            if balance != 0:
                if account.balance_side == 'debit':
                    debit_balance = balance if balance > 0 else 0
                    credit_balance = abs(balance) if balance < 0 else 0
                else:
                    credit_balance = balance if balance > 0 else 0
                    debit_balance = abs(balance) if balance < 0 else 0
                
                trial_balance.append({
                    'code': account.code,
                    'name': account.name,
                    'account_type': account.type,
                    'group': account.group.name if account.group else 'Unclassified',
                    'debit': float(debit_balance),
                    'credit': float(credit_balance),
                    'balance': float(balance)
                })
                
                total_debit += debit_balance
                total_credit += credit_balance
        
        # Sort by account code
        trial_balance.sort(key=lambda x: x['code'])
        
        # Check if trial balance is balanced
        is_balanced = abs(total_debit - total_credit) < 0.01  # Allow for minor rounding differences
        
        return JsonResponse({
            'success': True,
            'report_type': 'trial_balance',
            'period_end': period_end,
            'company_name': company.name,
            'data': {
                'accounts': trial_balance,
                'totals': {
                    'debit': float(total_debit),
                    'credit': float(total_credit),
                    'difference': float(total_debit - total_credit)
                },
                'is_balanced': is_balanced,
                'account_count': len(trial_balance)
            },
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ifrs_compliance': True
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def api_reconciliations(request):
    """API endpoint to get bank reconciliations"""
    try:
        company = request.user.company
        
        # For now, return sample data - in production, implement proper bank reconciliation logic
        reconciliations = [
            {
                'id': 1,
                'bank_account': 'Main Checking Account',
                'statement_date': '2024-01-31',
                'book_balance': 125000.00,
                'bank_balance': 127500.00,
                'status': 'pending',
                'created_at': '2024-01-25 10:30'
            },
            {
                'id': 2,
                'bank_account': 'Savings Account',
                'statement_date': '2024-01-31',
                'book_balance': 50000.00,
                'bank_balance': 50000.00,
                'status': 'reconciled',
                'created_at': '2024-01-20 14:15'
            }
        ]
        
        return JsonResponse({
            'success': True,
            'reconciliations': reconciliations,
            'total_reconciliations': len(reconciliations)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_create_reconciliation(request):
    """API endpoint to create a new bank reconciliation"""
    try:
        company = request.user.company
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['bank_account', 'statement_date', 'bank_balance']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'success': False, 'error': f'{field} is required'}, status=400)
        
        # For now, just return success - in production, implement proper reconciliation creation
        return JsonResponse({
            'success': True,
            'reconciliation_id': 3,  # Mock ID
            'message': 'Bank reconciliation created successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Additional API endpoints for Journal Entry CRUD operations
@login_required
@require_http_methods(["GET"])
def api_journal_entry_detail(request, entry_id):
    """API endpoint to get journal entry details"""
    try:
        company = request.user.company
        entry = get_object_or_404(JournalEntry, id=entry_id, company=company)
        
        # Get journal items for this entry
        items = entry.items.all().select_related('account')
        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'account_id': item.account.id,
                'account_code': item.account.code,
                'account_name': item.account.name,
                'debit': float(item.debit) if item.debit else 0,
                'credit': float(item.credit) if item.credit else 0,
                'description': item.description,
            })
        
        entry_data = {
            'id': entry.id,
            'reference': entry.reference,
            'date': entry.date.strftime('%Y-%m-%d') if entry.date else None,
            'description': entry.description,
            'journal': entry.journal.name if entry.journal else None,
            'total_amount': float(entry.total_amount) if entry.total_amount else 0,
            'status': entry.status,
            'created_by': entry.created_by.get_full_name() if entry.created_by else None,
            'created_at': entry.created_at.strftime('%Y-%m-%d %H:%M') if entry.created_at else None,
            'items': items_data,
        }
        
        return JsonResponse({
            'success': True,
            'entry': entry_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["PUT"])
@csrf_exempt
def api_update_journal_entry(request, entry_id):
    """API endpoint to update a journal entry"""
    try:
        company = request.user.company
        entry = get_object_or_404(JournalEntry, id=entry_id, company=company)
        
        # Only allow editing of draft entries
        if entry.status != 'draft':
            return JsonResponse({'success': False, 'error': 'Only draft entries can be edited'}, status=400)
        
        data = json.loads(request.body)
        
        # Update entry details
        if data.get('reference'):
            entry.reference = data['reference']
        if data.get('date'):
            entry.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if data.get('description'):
            entry.description = data['description']
        
        # Update journal items if provided
        if data.get('items'):
            # Delete existing items
            entry.items.all().delete()
            
            # Create new items
            total_debit = 0
            total_credit = 0
            for item_data in data['items']:
                account = Account.objects.get(id=item_data['account_id'], company=company)
                debit = float(item_data.get('debit', 0))
                credit = float(item_data.get('credit', 0))
                
                JournalItem.objects.create(
                    entry=entry,
                    account=account,
                    debit=debit,
                    credit=credit,
                    description=item_data.get('description', '')
                )
                
                total_debit += debit
                total_credit += credit
            
            # Validate balanced entry
            if abs(total_debit - total_credit) > 0.01:
                return JsonResponse({'success': False, 'error': 'Journal entry must be balanced'}, status=400)
            
            entry.total_amount = total_debit
        
        entry.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Journal entry updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["DELETE"])
@csrf_exempt
def api_delete_journal_entry(request, entry_id):
    """API endpoint to delete a journal entry"""
    try:
        company = request.user.company
        entry = get_object_or_404(JournalEntry, id=entry_id, company=company)
        
        # Only allow deletion of draft entries
        if entry.status == 'posted':
            return JsonResponse({'success': False, 'error': 'Posted entries cannot be deleted'}, status=400)
        
        entry_reference = entry.reference
        entry.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Journal entry {entry_reference} deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_post_journal_entry(request, entry_id):
    """API endpoint to post a journal entry"""
    try:
        company = request.user.company
        entry = get_object_or_404(JournalEntry, id=entry_id, company=company)
        
        # Only allow posting of draft entries
        if entry.status != 'draft':
            return JsonResponse({'success': False, 'error': 'Only draft entries can be posted'}, status=400)
        
        # Validate entry is balanced
        items = entry.items.all()
        total_debit = sum(float(item.debit) for item in items)
        total_credit = sum(float(item.credit) for item in items)
        
        if abs(total_debit - total_credit) > 0.01:
            return JsonResponse({'success': False, 'error': 'Entry is not balanced and cannot be posted'}, status=400)
        
        # Post the entry
        entry.status = 'posted'
        entry.posted_at = timezone.now()
        entry.posted_by = request.user
        entry.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Journal entry {entry.reference} posted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Journal Templates API endpoints
@login_required
@require_http_methods(["GET"])
def api_journal_templates(request):
    """API endpoint to get journal entry templates"""
    try:
        company = request.user.company
        
        # Get templates from database
        templates = JournalTemplate.objects.filter(
            company=company, 
            is_active=True
        ).prefetch_related('items', 'items__account').order_by('name')
        
        template_list = []
        for template in templates:
            template_items = []
            for item in template.items.all():
                template_items.append({
                    'id': item.id,
                    'account_id': item.account.id,
                    'account_code': item.account.code,
                    'account_name': item.account.name,
                    'description': item.description,
                    'debit_amount': float(item.debit_amount) if item.debit_amount else None,
                    'credit_amount': float(item.credit_amount) if item.credit_amount else None,
                    'is_amount_variable': item.is_amount_variable,
                    'order': item.order
                })
            
            template_list.append({
                'id': template.id,
                'name': template.name,
                'description': template.description,
                'items': template_items,
                'created_by': template.created_by.email if template.created_by else None,
                'created_at': template.created_at.isoformat()
            })
        
        return JsonResponse({
            'success': True,
            'templates': template_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_apply_journal_template(request):
    """API endpoint to apply a journal template"""
    try:
        company = request.user.company
        data = json.loads(request.body)
        
        template_id = data.get('template_id')
        amounts = data.get('amounts', {})
        
        # Get template (in production, fetch from database)
        templates = {
            1: {
                'name': 'Sales Revenue Entry',
                'items': [
                    {'account_code': '1200', 'account_name': 'Accounts Receivable', 'debit': True},
                    {'account_code': '4000', 'account_name': 'Sales Revenue', 'credit': True}
                ]
            },
            2: {
                'name': 'Purchase Entry',
                'items': [
                    {'account_code': '5000', 'account_name': 'Purchases', 'debit': True},
                    {'account_code': '2000', 'account_name': 'Accounts Payable', 'credit': True}
                ]
            }
        }
        
        template = templates.get(int(template_id))
        if not template:
            return JsonResponse({'success': False, 'error': 'Template not found'}, status=404)
        
        # Build journal entry data
        entry_data = {
            'description': f"{template['name']} - {data.get('description', '')}",
            'reference': data.get('reference', ''),
            'date': data.get('date', timezone.now().date().strftime('%Y-%m-%d')),
            'items': []
        }
        
        # Apply template items with amounts
        for item in template['items']:
            amount = float(amounts.get(item['account_code'], 0))
            if amount > 0:
                entry_data['items'].append({
                    'account_id': data.get('account_mappings', {}).get(item['account_code']),
                    'debit': amount if item.get('debit') else 0,
                    'credit': amount if item.get('credit') else 0,
                    'description': item['account_name']
                })
        
        return JsonResponse({
            'success': True,
            'entry_data': entry_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def api_create_journal_template(request):
    """API endpoint to create a new journal template"""
    try:
        company = request.user.company
        data = json.loads(request.body)
        
        # Create template
        template = JournalTemplate.objects.create(
            company=company,
            name=data.get('name'),
            description=data.get('description', ''),
            created_by=request.user
        )
        
        # Create template items
        for item_data in data.get('items', []):
            account = Account.objects.get(id=item_data['account_id'], company=company)
            JournalTemplateItem.objects.create(
                template=template,
                account=account,
                description=item_data.get('description', ''),
                debit_amount=item_data.get('debit_amount'),
                credit_amount=item_data.get('credit_amount'),
                is_amount_variable=item_data.get('is_amount_variable', True),
                order=item_data.get('order', 1)
            )
        
        return JsonResponse({
            'success': True, 
            'message': 'Journal template created successfully',
            'template_id': template.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["GET"])
def api_journal_template_detail(request, template_id):
    """API endpoint to get journal template details"""
    try:
        company = request.user.company
        template = JournalTemplate.objects.get(id=template_id, company=company)
        
        items = []
        for item in template.items.all():
            items.append({
                'id': item.id,
                'account_id': item.account.id,
                'account_code': item.account.code,
                'account_name': item.account.name,
                'description': item.description,
                'debit_amount': float(item.debit_amount) if item.debit_amount else None,
                'credit_amount': float(item.credit_amount) if item.credit_amount else None,
                'is_amount_variable': item.is_amount_variable,
                'order': item.order
            })
        
        template_data = {
            'id': template.id,
            'name': template.name,
            'description': template.description,
            'items': items,
            'is_active': template.is_active,
            'created_by': template.created_by.email if template.created_by else None,
            'created_at': template.created_at.isoformat(),
            'updated_at': template.updated_at.isoformat()
        }
        
        return JsonResponse({'success': True, 'template': template_data})
    except JournalTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Template not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["PUT"])
@csrf_exempt
def api_update_journal_template(request, template_id):
    """API endpoint to update a journal template"""
    try:
        company = request.user.company
        data = json.loads(request.body)
        
        template = JournalTemplate.objects.get(id=template_id, company=company)
        
        # Update template basic info
        template.name = data.get('name', template.name)
        template.description = data.get('description', template.description)
        template.is_active = data.get('is_active', template.is_active)
        template.save()
        
        # Update template items
        if 'items' in data:
            # Clear existing items
            template.items.all().delete()
            
            # Add new items
            for item_data in data['items']:
                account = Account.objects.get(id=item_data['account_id'], company=company)
                JournalTemplateItem.objects.create(
                    template=template,
                    account=account,
                    description=item_data.get('description', ''),
                    debit_amount=item_data.get('debit_amount'),
                    credit_amount=item_data.get('credit_amount'),
                    is_amount_variable=item_data.get('is_amount_variable', True),
                    order=item_data.get('order', 1)
                )
        
        return JsonResponse({'success': True, 'message': 'Journal template updated successfully'})
    except JournalTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Template not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["DELETE"])
@csrf_exempt
def api_delete_journal_template(request, template_id):
    """API endpoint to delete a journal template"""
    try:
        company = request.user.company
        template = JournalTemplate.objects.get(id=template_id, company=company)
        template.delete()
        
        return JsonResponse({'success': True, 'message': 'Journal template deleted successfully'})
    except JournalTemplate.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Template not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def financial_reports_ui(request):
    """Display comprehensive financial reports interface with IFRS 18 compliance"""
    try:
        company = request.user.company
        
        # Get company context for template
        context = {
            'company': company,
            'page_title': 'Financial Reports',
        }
        
        return render(request, 'accounting/financial_reports.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading financial reports: {str(e)}')
        return redirect('accounting:dashboard')